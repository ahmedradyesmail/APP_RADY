"""Temporary plate-only storage for صفحة التشيك (30-minute TTL)."""

from __future__ import annotations

import io
import json
import threading
import uuid
from contextlib import contextmanager
from datetime import datetime
from typing import Any

import openpyxl
import psycopg
from psycopg.rows import dict_row

from services.excel_utils import find_best_sheet, load_workbook_maybe_encrypted
from services.plate_utils import (
    auto_detect_plate_col,
    auto_detect_plate_col_from_row3,
    normalize_plate,
)

CHECK_TEMP_MAX_LARGE_BYTES = 15 * 1024 * 1024
CHECK_TEMP_TTL_MINUTES = 30

_schema_lock = threading.Lock()
_schema_ready: set[str] = set()


@contextmanager
def _tx(dsn: str, user_id: int, is_admin: bool):
    with psycopg.connect(dsn, autocommit=False) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT set_config('app.user_id', %s, true)", (str(int(user_id)),))
            cur.execute(
                "SELECT set_config('app.is_admin', %s, true)",
                ("true" if is_admin else "false",),
            )
        try:
            yield conn
            conn.commit()
        except Exception:
            conn.rollback()
            raise


def ensure_check_temp_schema(dsn: str) -> None:
    with _schema_lock:
        if dsn in _schema_ready:
            return
    with psycopg.connect(dsn, autocommit=True) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS check_temp_sessions (
                id BIGSERIAL PRIMARY KEY,
                token UUID NOT NULL UNIQUE,
                user_id INTEGER NOT NULL,
                created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
                last_seen_at TIMESTAMPTZ NOT NULL DEFAULT now()
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS check_temp_plates (
                id BIGSERIAL PRIMARY KEY,
                session_id BIGINT NOT NULL REFERENCES check_temp_sessions(id) ON DELETE CASCADE,
                user_id INTEGER NOT NULL,
                plate_normalized TEXT NOT NULL,
                created_at TIMESTAMPTZ NOT NULL DEFAULT now()
            )
            """
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_ctp_user_plate ON check_temp_plates (user_id, plate_normalized)"
        )
        conn.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_ctp_session_plate ON check_temp_plates (session_id, plate_normalized)"
        )
        conn.execute("ALTER TABLE check_temp_sessions ENABLE ROW LEVEL SECURITY")
        conn.execute("ALTER TABLE check_temp_sessions FORCE ROW LEVEL SECURITY")
        conn.execute("ALTER TABLE check_temp_plates ENABLE ROW LEVEL SECURITY")
        conn.execute("ALTER TABLE check_temp_plates FORCE ROW LEVEL SECURITY")
        conn.execute(
            """
            DO $p$
            BEGIN
              IF NOT EXISTS (
                SELECT 1 FROM pg_policies
                WHERE schemaname='public' AND tablename='check_temp_sessions'
                  AND policyname='check_temp_sessions_isolation'
              ) THEN
                CREATE POLICY check_temp_sessions_isolation ON check_temp_sessions FOR ALL
                USING (
                  COALESCE(current_setting('app.is_admin', true), '') = 'true'
                  OR user_id = (NULLIF(current_setting('app.user_id', true), ''))::integer
                )
                WITH CHECK (
                  COALESCE(current_setting('app.is_admin', true), '') = 'true'
                  OR user_id = (NULLIF(current_setting('app.user_id', true), ''))::integer
                );
              END IF;
            END
            $p$;
            """
        )
        conn.execute(
            """
            DO $p$
            BEGIN
              IF NOT EXISTS (
                SELECT 1 FROM pg_policies
                WHERE schemaname='public' AND tablename='check_temp_plates'
                  AND policyname='check_temp_plates_isolation'
              ) THEN
                CREATE POLICY check_temp_plates_isolation ON check_temp_plates FOR ALL
                USING (
                  COALESCE(current_setting('app.is_admin', true), '') = 'true'
                  OR user_id = (NULLIF(current_setting('app.user_id', true), ''))::integer
                )
                WITH CHECK (
                  COALESCE(current_setting('app.is_admin', true), '') = 'true'
                  OR user_id = (NULLIF(current_setting('app.user_id', true), ''))::integer
                );
              END IF;
            END
            $p$;
            """
        )
    with _schema_lock:
        _schema_ready.add(dsn)


def purge_expired_temp_sessions_sync(dsn: str) -> int:
    ensure_check_temp_schema(dsn)
    with psycopg.connect(dsn, autocommit=True) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                DELETE FROM check_temp_sessions
                WHERE last_seen_at < (now() - make_interval(mins => %s))
                """,
                (CHECK_TEMP_TTL_MINUTES,),
            )
            return int(cur.rowcount or 0)


def start_temp_session_sync(dsn: str, user_id: int, is_admin: bool) -> str:
    ensure_check_temp_schema(dsn)
    tok = str(uuid.uuid4())
    with _tx(dsn, user_id, is_admin) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO check_temp_sessions (token, user_id)
                VALUES (%s::uuid, %s)
                """,
                (tok, user_id),
            )
    return tok


def ping_temp_session_sync(dsn: str, user_id: int, is_admin: bool, session_token: str) -> bool:
    ensure_check_temp_schema(dsn)
    with _tx(dsn, user_id, is_admin) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE check_temp_sessions
                SET last_seen_at = now()
                WHERE token = %s::uuid AND user_id = %s
                """,
                (session_token, user_id),
            )
            return cur.rowcount > 0


def _get_session_id(cur, user_id: int, session_token: str) -> int | None:
    cur.execute(
        "SELECT id FROM check_temp_sessions WHERE token = %s::uuid AND user_id = %s",
        (session_token, user_id),
    )
    row = cur.fetchone()
    return int(row[0]) if row else None


def delete_temp_session_sync(
    dsn: str, user_id: int, is_admin: bool, session_token: str
) -> bool:
    ensure_check_temp_schema(dsn)
    tok = (session_token or "").strip()
    if not tok:
        return False
    with _tx(dsn, user_id, is_admin) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                DELETE FROM check_temp_sessions
                WHERE token = %s::uuid AND user_id = %s
                """,
                (tok, user_id),
            )
            return cur.rowcount > 0


def temp_plate_exists_sync(
    dsn: str,
    user_id: int,
    is_admin: bool,
    *,
    session_token: str,
    plate_text: str,
) -> bool:
    ensure_check_temp_schema(dsn)
    pn = normalize_plate(plate_text or "")
    if not pn:
        return False
    with _tx(dsn, user_id, is_admin) as conn:
        with conn.cursor() as cur:
            sid = _get_session_id(cur, user_id, session_token)
            if sid is None:
                return False
            cur.execute("UPDATE check_temp_sessions SET last_seen_at = now() WHERE id = %s", (sid,))
            cur.execute(
                """
                SELECT 1
                FROM check_temp_plates
                WHERE session_id = %s AND user_id = %s AND plate_normalized = %s
                LIMIT 1
                """,
                (sid, user_id, pn),
            )
            return cur.fetchone() is not None


def upload_large_temp_plates_sync(
    dsn: str,
    user_id: int,
    is_admin: bool,
    *,
    session_token: str,
    large_bytes: bytes,
    password: str,
    large_col: str,
    large_sheet: str,
) -> dict[str, Any]:
    ensure_check_temp_schema(dsn)
    if len(large_bytes) > CHECK_TEMP_MAX_LARGE_BYTES:
        raise ValueError("الملف الكبير يتجاوز 15 ميجابايت")

    wb = load_workbook_maybe_encrypted(large_bytes, password)
    try:
        ws = wb[large_sheet] if large_sheet and large_sheet in wb.sheetnames else find_best_sheet(wb)
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError("الملف الكبير فارغ")
        headers = [str(h).strip() if h is not None else "" for h in header]
        row2 = next(rows, None)
        row3 = next(rows, None)
        detected = auto_detect_plate_col(headers) or auto_detect_plate_col_from_row3(headers, row3)
        col = large_col.strip() or (detected or "")
        if not col or col not in headers:
            raise ValueError(f"لم يُعثر على عمود اللوحة في الملف الكبير. الأعمدة: {headers}")
        ci = headers.index(col)

        all_rows = []
        if row2 is not None:
            all_rows.append(row2)
        if row3 is not None:
            all_rows.append(row3)
        all_rows.extend(list(rows))

        values: list[tuple[int, int, str]] = []
        for row in all_rows:
            if all(v is None for v in row):
                continue
            raw = row[ci] if ci < len(row) else None
            pn = normalize_plate(raw)
            if pn:
                values.append((0, user_id, pn))
        if not values:
            raise ValueError("لا توجد لوحات صالحة في الملف الكبير")

        with _tx(dsn, user_id, is_admin) as conn:
            with conn.cursor() as cur:
                sid = _get_session_id(cur, user_id, session_token)
                if sid is None:
                    raise ValueError("جلسة التشيك غير صالحة أو انتهت")
                cur.execute(
                    "UPDATE check_temp_sessions SET last_seen_at = now() WHERE id = %s",
                    (sid,),
                )
                cur.execute("DELETE FROM check_temp_plates WHERE session_id = %s", (sid,))
                values = [(sid, user_id, p) for (_, user_id, p) in values]
                cur.executemany(
                    """
                    INSERT INTO check_temp_plates (session_id, user_id, plate_normalized)
                    VALUES (%s, %s, %s)
                    ON CONFLICT DO NOTHING
                    """,
                    values,
                )
                cur.execute(
                    "SELECT COUNT(*)::int FROM check_temp_plates WHERE session_id = %s",
                    (sid,),
                )
                count = int(cur.fetchone()[0])
        return {
            "plate_column_used": col,
            "stored_count": count,
            "sheet_name": ws.title,
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass


def query_temp_plates_sync(
    dsn: str,
    user_id: int,
    is_admin: bool,
    *,
    session_token: str,
    plates_text: str,
) -> dict[str, Any]:
    ensure_check_temp_schema(dsn)
    inp = []
    for line in (plates_text or "").splitlines():
        v = " ".join(str(line).strip().split())
        if v:
            inp.append(v)
    if not inp:
        raise ValueError("لا توجد لوحات نصية للفحص")

    rows: list[dict[str, Any]] = []
    with _tx(dsn, user_id, is_admin) as conn:
        with conn.cursor(row_factory=dict_row) as cur:
            sid = _get_session_id(cur, user_id, session_token)
            if sid is None:
                raise ValueError("جلسة التشيك غير صالحة أو انتهت")
            cur.execute("UPDATE check_temp_sessions SET last_seen_at = now() WHERE id = %s", (sid,))
            for p in inp:
                pn = normalize_plate(p)
                exists = False
                if pn:
                    cur.execute(
                        """
                        SELECT 1
                        FROM check_temp_plates
                        WHERE session_id = %s AND user_id = %s AND plate_normalized = %s
                        LIMIT 1
                        """,
                        (sid, user_id, pn),
                    )
                    exists = cur.fetchone() is not None
                rows.append({"plate": p, "exists": bool(exists)})
    found = sum(1 for r in rows if r["exists"])
    return {"rows": rows, "total": len(rows), "found": found, "not_found": len(rows) - found}
