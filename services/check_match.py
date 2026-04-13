"""Synchronous plate check logic (runs in thread pool / background)."""

from __future__ import annotations

import io
import itertools
import json
import logging
import os
import re
import sqlite3
import tempfile
from datetime import datetime

import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from services.plate_utils import (
    normalize_plate,
    auto_detect_plate_col,
    auto_detect_plate_col_from_row3,
)
from services.excel_utils import (
    load_workbook_maybe_encrypted,
    find_best_sheet,
    workbook_to_bytes,
)

logger = logging.getLogger(__name__)

PREVIEW_MAX_ROWS = 300

_SMALL_HDR_PATTERNS: tuple[tuple[re.Pattern[str], str], ...] = (
    (re.compile(r"^\s*صغير\s*[—–\-]\s*", re.IGNORECASE), ""),
    (re.compile(r"^\s*صغير\s+", re.IGNORECASE), ""),
    (re.compile(r"\s+صغير\s*$", re.IGNORECASE), ""),
    (re.compile(r"\s*[—–\-]\s*صغير\s*$", re.IGNORECASE), ""),
)


def _strip_small_word_from_header_title(title: str) -> str:
    """Remove «صغير» and common prefixes from header text in the exported sheet."""
    raw = (title or "").strip()
    if not raw:
        return raw
    s = raw
    for pat, rep in _SMALL_HDR_PATTERNS:
        s = pat.sub(rep, s).strip()
    s = re.sub(r"\s*صغير\s*", " ", s)
    s = " ".join(s.split()).strip()
    return s if s else raw


def _close_wb(wb) -> None:
    if wb is None:
        return
    try:
        wb.close()
    except Exception:
        pass


def _norm_large_export_cols(requested: list[str], available: list[str]) -> list[str]:
    """Empty selection → export all large columns."""
    if not requested:
        return [h for h in available if h]
    out = [c for c in requested if c in available]
    return out if out else [h for h in available if h]


def _norm_small_export_cols(requested: list[str], available: list[str]) -> list[str]:
    """Empty selection → no small columns (user must tick columns explicitly)."""
    if not requested:
        return []
    return [c for c in requested if c in available]


def _sqlite_connect(path: str) -> sqlite3.Connection:
    con = sqlite3.connect(path)
    con.execute("PRAGMA journal_mode=OFF")
    con.execute("PRAGMA synchronous=OFF")
    con.execute("PRAGMA temp_store=MEMORY")
    con.execute("PRAGMA cache_size=-20000")
    return con


def _sqlite_init_index(con: sqlite3.Connection) -> None:
    con.execute(
        "CREATE TABLE IF NOT EXISTS plate_idx ("
        "plate_key TEXT NOT NULL, "
        "payload_json TEXT NOT NULL)"
    )
    con.execute("CREATE INDEX IF NOT EXISTS idx_plate_key ON plate_idx(plate_key)")
    con.commit()


def _sqlite_insert_batch(
    con: sqlite3.Connection, rows: list[tuple[str, str]]
) -> None:
    if not rows:
        return
    con.executemany(
        "INSERT INTO plate_idx(plate_key, payload_json) VALUES(?,?)",
        rows,
    )
    con.commit()


def _make_border() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _build_styled_row(
    ws,
    values: list,
    col_sources: list[str],
    header: bool,
    header_font,
    body_font,
    header_fill_large,
    header_fill_small,
    body_fill_large,
    body_fill_small,
    align_header,
    align_cell,
    border,
) -> None:
    row_cells = []
    for i, v in enumerate(values):
        src = col_sources[i] if i < len(col_sources) else "large"
        c = WriteOnlyCell(ws, value=v)
        c.border = border
        if header:
            c.font = header_font
            c.fill = header_fill_large if src == "large" else header_fill_small
            c.alignment = align_header
        else:
            c.font = body_font
            c.fill = body_fill_large if src == "large" else body_fill_small
            c.alignment = align_cell
        row_cells.append(c)
    ws.append(row_cells)


def run_check_plates_sync(
    lc_bytes: bytes,
    sc_bytes: bytes,
    password: str,
    large_col: str,
    small_col: str,
    large_sheet: str,
    small_sheet: str,
    large_export_cols: list[str] | None = None,
    small_export_cols: list[str] | None = None,
) -> dict:
    """
    Returns one of:
      {"kind": "xlsx", "content": bytes, "filename": str}
      {"kind": "json", "status_code": int, "body": dict}
    """
    large_wb = None
    small_wb = None
    sqlite_con = None
    sqlite_path = ""
    try:
        try:
            large_wb = load_workbook_maybe_encrypted(lc_bytes, password)
        except ValueError:
            logger.exception("Failed to open encrypted large workbook")
            raise

        try:
            small_wb = openpyxl.load_workbook(
                io.BytesIO(sc_bytes), read_only=True, data_only=True
            )
        except Exception as e:
            raise ValueError(f"تعذّر فتح الملف الصغير: {e}") from e

        large_ws = (
            large_wb[large_sheet]
            if large_sheet and large_sheet in large_wb.sheetnames
            else find_best_sheet(large_wb)
        )
        small_ws = (
            small_wb[small_sheet]
            if small_sheet and small_sheet in small_wb.sheetnames
            else find_best_sheet(small_wb)
        )

        lrows = large_ws.iter_rows(values_only=True)
        header_l = next(lrows, None)
        if header_l is None:
            raise ValueError("الملف الكبير فارغ")

        lh = [str(h).strip() if h is not None else "" for h in header_l]
        lc = large_col.strip() or auto_detect_plate_col(lh)

        if not lc or lc not in lh:
            return {
                "kind": "json",
                "status_code": 422,
                "body": {
                    "detail": (
                        f"لم يُعثر على عمود اللوحة في الملف الكبير "
                        f"(شيت: {large_ws.title}). الأعمدة: {lh}"
                    ),
                    "headers": lh,
                    "code": "COL_NOT_FOUND_LARGE",
                },
            }

        le_cols = _norm_large_export_cols(list(large_export_cols or []), lh)
        le_idx = [lh.index(c) for c in le_cols]
        lci = lh.index(lc)
        tmp = tempfile.NamedTemporaryFile(prefix="check_idx_", suffix=".sqlite3", delete=False)
        sqlite_path = tmp.name
        tmp.close()
        sqlite_con = _sqlite_connect(sqlite_path)
        _sqlite_init_index(sqlite_con)
        batch: list[tuple[str, str]] = []
        batch_size = 2000
        for row in lrows:
            if all(v is None for v in row):
                continue
            rp = row[lci] if lci < len(row) else None
            norm = normalize_plate(rp)
            if not norm:
                continue
            out_vals = [(row[i] if i < len(row) else None) for i in le_idx]
            batch.append((norm, json.dumps(out_vals, ensure_ascii=False)))
            if len(batch) >= batch_size:
                _sqlite_insert_batch(sqlite_con, batch)
                batch.clear()
        if batch:
            _sqlite_insert_batch(sqlite_con, batch)

        srows = small_ws.iter_rows(values_only=True)
        header_s = next(srows, None)
        if header_s is None:
            raise ValueError("الملف الصغير فارغ")

        sh = [str(h).strip() if h is not None else "" for h in header_s]
        row2 = next(srows, None)
        row3 = next(srows, None)
        detected_small = auto_detect_plate_col(sh) or auto_detect_plate_col_from_row3(sh, row3)
        sc = small_col.strip() or detected_small

        if not sc or sc not in sh:
            return {
                "kind": "json",
                "status_code": 422,
                "body": {
                    "detail": f"لم يُعثر على عمود اللوحة في الملف الصغير. الأعمدة: {sh}",
                    "headers": sh,
                    "code": "COL_NOT_FOUND_SMALL",
                },
            }

        se_cols = _norm_small_export_cols(list(small_export_cols or []), sh)
        se_idx = [sh.index(c) for c in se_cols]
        sci = sh.index(sc)
        matched_rows_count = 0
        matched_plate_hits = 0
        unmatched_plates = 0

        display_headers: list[str] = [
            _strip_small_word_from_header_title(c) for c in le_cols
        ] + [_strip_small_word_from_header_title(c) for c in se_cols]
        col_sources: list[str] = ["large"] * len(le_cols) + ["small"] * len(se_cols)

        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
        body_font = Font(name="Arial", size=11, color="000000")
        header_fill_large = PatternFill("solid", start_color="1E40AF")
        header_fill_small = PatternFill("solid", start_color="166534")
        body_fill_large = PatternFill("solid", start_color="DBEAFE")
        body_fill_small = PatternFill("solid", start_color="DCFCE7")
        align_header = Alignment(horizontal="center", vertical="center")
        align_cell = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = _make_border()

        wb_out = openpyxl.Workbook(write_only=True)
        ws_m = wb_out.create_sheet(title="التطابقات")
        ws_m.sheet_view.rightToLeft = True
        _build_styled_row(
            ws_m,
            display_headers,
            col_sources,
            True,
            header_font,
            body_font,
            header_fill_large,
            header_fill_small,
            body_fill_large,
            body_fill_small,
            align_header,
            align_cell,
            border,
        )

        preview_rows: list[list[str]] = []

        for row in itertools.chain((r for r in (row2, row3) if r is not None), srows):
            if all(v is None for v in row):
                continue
            rp = row[sci] if sci < len(row) else None
            norm = normalize_plate(rp)
            if not norm:
                continue
            small_vals = [(row[i] if i < len(row) else None) for i in se_idx]
            large_candidates = []
            cur = sqlite_con.execute(
                "SELECT payload_json FROM plate_idx WHERE plate_key = ?",
                (norm,),
            )
            for (payload_json,) in cur:
                try:
                    large_candidates.append(json.loads(payload_json))
                except Exception:
                    continue
            if large_candidates:
                matched_plate_hits += 1
                for large_vals in large_candidates:
                    row_out = [*large_vals, *small_vals]
                    _build_styled_row(
                        ws_m,
                        row_out,
                        col_sources,
                        False,
                        header_font,
                        body_font,
                        header_fill_large,
                        header_fill_small,
                        body_fill_large,
                        body_fill_small,
                        align_header,
                        align_cell,
                        border,
                    )
                    matched_rows_count += 1
                    if len(preview_rows) < PREVIEW_MAX_ROWS:
                        preview_rows.append(
                            ["" if v is None else str(v).strip() for v in row_out]
                        )
            else:
                unmatched_plates += 1

        if not matched_rows_count:
            return {
                "kind": "json",
                "status_code": 200,
                "body": {
                    "detail": "لا توجد تطابقات بين الملفين",
                    "matched": 0,
                    "unmatched": unmatched_plates,
                    "large_col_used": lc,
                    "small_col_used": sc,
                },
            }

        content = workbook_to_bytes(wb_out)
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"التطابقات_{ts}.xlsx"

        return {
            "kind": "xlsx",
            "content": content,
            "filename": filename,
            "preview": {
                "headers": display_headers,
                "col_sources": col_sources,
                "rows": preview_rows,
                "total_rows": matched_rows_count,
                "truncated": matched_rows_count > PREVIEW_MAX_ROWS,
                "stats": {
                    "matched_rows": matched_rows_count,
                    "matched_plate_hits": matched_plate_hits,
                    "unmatched_plates": unmatched_plates,
                },
            },
        }
    finally:
        try:
            if sqlite_con is not None:
                sqlite_con.close()
        except Exception:
            pass
        if sqlite_path:
            try:
                os.unlink(sqlite_path)
            except OSError:
                logger.debug("Could not delete sqlite temp index: %s", sqlite_path)
        _close_wb(small_wb)
        _close_wb(large_wb)
