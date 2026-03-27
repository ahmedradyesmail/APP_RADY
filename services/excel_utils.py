import asyncio
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .plate_utils import auto_detect_plate_col

# Project uses openpyxl (not pandas); heavy workbook I/O runs in a thread pool.


def load_workbook_maybe_encrypted(
    file_bytes: bytes, password: str = ""
) -> openpyxl.Workbook:
    """Load workbook, optionally decrypting with password."""
    if password:
        try:
            import msoffcrypto
            enc = io.BytesIO(file_bytes)
            dec = io.BytesIO()
            of = msoffcrypto.OfficeFile(enc)
            of.load_key(password=password)
            of.decrypt(dec)
            dec.seek(0)
            return openpyxl.load_workbook(dec, read_only=True, data_only=True)
        except Exception as e:
            raise ValueError(f"فشل فك تشفير الملف — تحقق من كلمة المرور: {e}")
    try:
        return openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
    except Exception as e:
        raise ValueError(f"تعذّر فتح الملف: {e}")


def find_best_sheet(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    """Find the sheet that has a plate column, otherwise the largest sheet."""
    best_ws = None
    best_n = -1
    for sn in wb.sheetnames:
        ws = wb[sn]
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(h).strip() if h is not None else "" for h in row]
            break
        if auto_detect_plate_col(headers):
            return ws
        try:
            n = ws.max_row or 0
        except Exception:
            n = 0
        if n > best_n:
            best_n = n
            best_ws = ws
    return best_ws if best_ws is not None else wb.active


def make_border() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_excel_style(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    headers: list[str],
    rows_of_dicts: list[dict],
    header_color: str = "1F4E79",
) -> None:
    """Apply professional RTL styling to a worksheet."""
    ws.sheet_view.rightToLeft = True

    hf    = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    hfill = PatternFill("solid", start_color=header_color)
    ha    = Alignment(horizontal="center", vertical="center")
    ca    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    df    = Font(name="Arial", size=11)
    brd   = make_border()
    fe    = PatternFill("solid", start_color="D6E4F0")
    fo    = PatternFill("solid", start_color="FFFFFF")

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = ha
        cell.border = brd
    ws.row_dimensions[1].height = 30

    for ri, rd in enumerate(rows_of_dicts, 1):
        fill = fe if ri % 2 == 0 else fo
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=ri + 1, column=ci, value=rd.get(h, ""))
            cell.font = df
            cell.alignment = ca
            cell.border = brd
            cell.fill = fill

    for cc in ws.columns:
        try:
            length = max(len(str(c.value or "")) for c in cc)
            ws.column_dimensions[cc[0].column_letter].width = min(length + 4, 40)
        except Exception:
            pass


def workbook_to_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def load_workbook_maybe_encrypted_async(
    file_bytes: bytes, password: str = ""
) -> openpyxl.Workbook:
    return await asyncio.to_thread(
        load_workbook_maybe_encrypted, file_bytes, password
    )


async def find_best_sheet_async(wb: openpyxl.Workbook):
    return await asyncio.to_thread(find_best_sheet, wb)


async def workbook_to_bytes_async(wb: openpyxl.Workbook) -> bytes:
    return await asyncio.to_thread(workbook_to_bytes, wb)


async def load_workbook_from_bytes_async(
    content: bytes, read_only: bool = True, data_only: bool = True
) -> openpyxl.Workbook:
    def _load() -> openpyxl.Workbook:
        return openpyxl.load_workbook(
            io.BytesIO(content), read_only=read_only, data_only=data_only
        )

    return await asyncio.to_thread(_load)