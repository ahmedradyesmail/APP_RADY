"""Postgres-backed storage for الفرز: عدة ملفات كبيرة لكل مستخدم + RLS."""

from __future__ import annotations

import csv
import io
import itertools
import json
import logging
import os
import re
import unicodedata
import tempfile
import threading
from contextlib import contextmanager
from datetime import datetime
from typing import Any

import openpyxl
import psycopg
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from psycopg.rows import dict_row
from psycopg.types.json import Json

from services.check_match import (
    PREVIEW_MAX_ROWS,
    _norm_large_export_cols,
    _norm_small_export_cols,
    _strip_small_word_from_header_title,
)
from services.excel_utils import find_best_sheet, load_workbook_maybe_encrypted, workbook_to_bytes
from services.plate_utils import (
    auto_detect_plate_col,
    auto_detect_plate_col_from_row3,
    normalize_plate,
)

logger = logging.getLogger(__name__)

CHECK_PG_MAX_LARGE_BYTES = 15 * 1024 * 1024
CHECK_PG_MAX_ROWS_PER_USER = 3_000_000
GPS_HEADER = "GPS"
# Max distinct plates per SQL round-trip (ANY(...)); avoids N queries per small-file row.
_CHECK_PLATE_LOOKUP_BATCH = 2048


def _norm_header_sim(s: str) -> str:
    """
    Same idea as static/_om_field.js omNormHeaderForSim: map union export names
    to per-file JSONB keys (e.g. اللوحة vs اللوحه, أ vs ا).
    """
    t = unicodedata.normalize("NFKC", str(s or ""))
    for ch in ("\u0640", "\u061c", "\u200c", "\u200d", "\u200e", "\u200f", "\ufeff"):
        t = t.replace(ch, "")
    for ch in ("\u0623", "\u0625", "\u0622"):
        t = t.replace(ch, "\u0627")
    t = t.replace("\u0629", "\u0647").replace("\u0649", "\u064a")
    t = re.sub(r"\s+", "", t).strip().lower()
    return t


def _row_norm_key_index(row_dict: dict[str, Any]) -> dict[str, str]:
    """normalized header -> first original key in row_data."""
    idx: dict[str, str] = {}
    for k in row_dict:
        ks = str(k) if k is not None else ""
        nk = _norm_header_sim(ks)
        if nk and nk not in idx:
            idx[nk] = ks
    return idx


def _map_export_headers_to_sheet(requested: list[str], lh: list[str]) -> list[str]:
    """
    أعمدة الاختيار في الواجهة تُبنى من اتحاد كل الملفات؛ نربط كل اسم مختار
    بعمود الشيت الحقيقي (نفس ترتيب الاختيار) حتى تطابق مفاتيح JSONB.
    """
    lh_list = [str(h).strip() for h in lh if h and str(h).strip()]
    req_clean = [str(c).strip() for c in (requested or []) if c and str(c).strip()]
    if not lh_list:
        # استيراد قديم / column_order فاضي في DB — لا نُسقط أعمدة الملف الكبير
        return req_clean
    if not requested:
        return lh_list
    lh_set = set(lh_list)
    norm_to_lh: dict[str, str] = {}
    for h in lh_list:
        nk = _norm_header_sim(h)
        if nk and nk not in norm_to_lh:
            norm_to_lh[nk] = h
    out: list[str] = []
    used_nk: set[str] = set()
    for c in requested:
        cs = str(c).strip() if c else ""
        if not cs:
            continue
        if cs in lh_set:
            nk = _norm_header_sim(cs)
            if nk not in used_nk:
                used_nk.add(nk)
                out.append(cs)
            continue
        nk = _norm_header_sim(cs)
        real = norm_to_lh.get(nk)
        if real and nk not in used_nk:
            used_nk.add(nk)
            out.append(real)
    return out if out else lh_list


# ألوان ترويسة أقسام الملفات الكبيرة (دوّار)
FILE_SECTION_HEADER_FILLS = [
    "4472C4",
    "C65911",
    "70AD47",
    "9F4F96",
    "FFC000",
    "E15759",
    "5B9BD5",
    "264478",
    "ED7D31",
    "A5A5A5",
]

_schema_lock = threading.Lock()
_schema_initialized: set[str] = set()


def _jsonable_cell(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    from datetime import date, time
    from datetime import datetime as dt

    if isinstance(v, (dt, date, time)):
        return str(v)
    return str(v)


def _safe_filename(name: str) -> str:
    s = (name or "").strip() or "upload.xlsx"
    s = re.sub(r"[/\\<>|\":?*]", "_", s)[:240]
    return s


@contextmanager
def check_pg_tx(dsn: str, user_id: int, is_admin: bool):
    with psycopg.connect(dsn, autocommit=False) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT set_config('app.user_id', %s, true)", (str(int(user_id)),))
            cur.execute(
                "SELECT set_config('app.is_admin', %s, true)",
                ("true" if is_admin else "false",),
            )
        try:
            yield conn
            conn.commit()
        except Exception:
            conn.rollback()
            raise


def _thin_border() -> Border:
    t = Side(style="thin", color="BFBFBF")
    return Border(left=t, right=t, top=t, bottom=t)


def _apply_migrations(conn) -> None:
    """Idempotent: check_large_imports + import_id على check_large_rows."""
    with conn.cursor() as cur:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS check_large_imports (
                id BIGSERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL,
                filename TEXT NOT NULL DEFAULT '',
                sheet_name TEXT NOT NULL DEFAULT '',
                plate_column TEXT NOT NULL DEFAULT '',
                column_order JSONB NOT NULL DEFAULT '[]'::jsonb,
                created_at TIMESTAMPTZ NOT NULL DEFAULT now()
            )
            """
        )
        cur.execute(
            """
            SELECT EXISTS (
              SELECT 1 FROM information_schema.tables
              WHERE table_schema = 'public' AND table_name = 'check_large_rows'
            )
            """
        )
        rows_exists = cur.fetchone()[0]
        if not rows_exists:
            cur.execute(
                """
                CREATE TABLE check_large_rows (
                    id BIGSERIAL PRIMARY KEY,
                    user_id INTEGER NOT NULL,
                    import_id BIGINT NOT NULL REFERENCES check_large_imports(id) ON DELETE CASCADE,
                    plate_normalized TEXT NOT NULL,
                    row_data JSONB NOT NULL,
                    gps TEXT NOT NULL DEFAULT ''
                )
                """
            )
        else:
            cur.execute(
                """
                SELECT EXISTS (
                  SELECT 1 FROM information_schema.columns
                  WHERE table_schema = 'public' AND table_name = 'check_large_rows'
                    AND column_name = 'import_id'
                )
                """
            )
            has_import_id = cur.fetchone()[0]
            if not has_import_id:
                cur.execute(
                    "ALTER TABLE check_large_rows ADD COLUMN import_id BIGINT REFERENCES check_large_imports(id) ON DELETE CASCADE"
                )
                cur.execute(
                    """
                    INSERT INTO check_large_imports (user_id, filename, sheet_name, plate_column, column_order, created_at)
                    SELECT m.user_id, 'legacy-import', COALESCE(m.sheet_name, ''), '', m.column_order, m.updated_at
                    FROM check_large_meta m
                    WHERE NOT EXISTS (
                      SELECT 1 FROM check_large_imports i WHERE i.user_id = m.user_id AND i.filename = 'legacy-import'
                    )
                    """
                )
                cur.execute(
                    """
                    UPDATE check_large_rows r
                    SET import_id = i.id
                    FROM check_large_imports i
                    WHERE r.import_id IS NULL AND i.user_id = r.user_id AND i.filename = 'legacy-import'
                    """
                )
                cur.execute(
                    """
                    INSERT INTO check_large_imports (user_id, filename, sheet_name, plate_column, column_order)
                    SELECT DISTINCT r.user_id, 'migrated', '', '', '[]'::jsonb
                    FROM check_large_rows r
                    WHERE r.import_id IS NULL
                      AND NOT EXISTS (
                        SELECT 1 FROM check_large_imports i
                        WHERE i.user_id = r.user_id AND i.filename = 'migrated'
                      )
                    """
                )
                cur.execute(
                    """
                    UPDATE check_large_rows r
                    SET import_id = i.id
                    FROM check_large_imports i
                    WHERE r.import_id IS NULL AND i.user_id = r.user_id AND i.filename = 'migrated'
                    """
                )
                cur.execute("DELETE FROM check_large_rows WHERE import_id IS NULL")
                cur.execute(
                    "ALTER TABLE check_large_rows ALTER COLUMN import_id SET NOT NULL"
                )
        cur.execute("DROP TABLE IF EXISTS check_large_meta CASCADE")
        cur.execute(
            "CREATE INDEX IF NOT EXISTS idx_clr_import_plate ON check_large_rows (import_id, plate_normalized)"
        )
        cur.execute(
            "CREATE INDEX IF NOT EXISTS idx_clr_user ON check_large_rows (user_id)"
        )
        cur.execute(
            "CREATE INDEX IF NOT EXISTS idx_clr_user_plate ON check_large_rows (user_id, plate_normalized)"
        )
        cur.execute(
            "CREATE INDEX IF NOT EXISTS idx_clr_plate_normalized ON check_large_rows (plate_normalized)"
        )


def _compact_redundant_gps_in_row_data(conn, batch_size: int = 20000, max_batches: int = 3) -> int:
    """
    Save space safely by removing duplicated GPS key from row_data only when
    dedicated `gps` column is already non-empty.
    """
    updated_total = 0
    with conn.cursor() as cur:
        for _ in range(max_batches):
            cur.execute(
                """
                WITH picked AS (
                    SELECT ctid
                    FROM check_large_rows
                    WHERE gps <> ''
                      AND (row_data ? 'GPS' OR row_data ? 'gps')
                    LIMIT %s
                )
                UPDATE check_large_rows r
                SET row_data = (r.row_data - 'GPS' - 'gps')
                FROM picked
                WHERE r.ctid = picked.ctid
                """,
                (batch_size,),
            )
            n = int(cur.rowcount or 0)
            if n <= 0:
                break
            updated_total += n
    if updated_total:
        logger.info("check_postgres compacted redundant GPS row_data rows=%s", updated_total)
    return updated_total


def _ensure_peer_group_mirror_and_rls(conn) -> None:
    """Shared فرز data: mirror table + RLS so group peers can read each other's rows."""
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS check_mirror_user_groups (
            user_id INTEGER NOT NULL PRIMARY KEY,
            group_id INTEGER NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_cmug_group ON check_mirror_user_groups (group_id)"
    )
    conn.execute("DROP POLICY IF EXISTS check_large_rows_isolation ON check_large_rows")
    conn.execute("DROP POLICY IF EXISTS check_large_imports_isolation ON check_large_imports")
    conn.execute(
        """
        CREATE POLICY check_large_rows_isolation ON check_large_rows FOR ALL
        USING (
          COALESCE(current_setting('app.is_admin', true), '') = 'true'
          OR user_id = (NULLIF(current_setting('app.user_id', true), ''))::integer
          OR EXISTS (
            SELECT 1 FROM check_mirror_user_groups m1
            INNER JOIN check_mirror_user_groups m2 ON m1.group_id = m2.group_id
            WHERE m1.user_id = (NULLIF(current_setting('app.user_id', true), ''))::integer
              AND m2.user_id = check_large_rows.user_id
          )
        )
        WITH CHECK (
          COALESCE(current_setting('app.is_admin', true), '') = 'true'
          OR user_id = (NULLIF(current_setting('app.user_id', true), ''))::integer
        )
        """
    )
    conn.execute(
        """
        CREATE POLICY check_large_imports_isolation ON check_large_imports FOR ALL
        USING (
          COALESCE(current_setting('app.is_admin', true), '') = 'true'
          OR user_id = (NULLIF(current_setting('app.user_id', true), ''))::integer
          OR EXISTS (
            SELECT 1 FROM check_mirror_user_groups m1
            INNER JOIN check_mirror_user_groups m2 ON m1.group_id = m2.group_id
            WHERE m1.user_id = (NULLIF(current_setting('app.user_id', true), ''))::integer
              AND m2.user_id = check_large_imports.user_id
          )
        )
        WITH CHECK (
          COALESCE(current_setting('app.is_admin', true), '') = 'true'
          OR user_id = (NULLIF(current_setting('app.user_id', true), ''))::integer
        )
        """
    )


def ensure_check_pg_schema(dsn: str) -> None:
    with _schema_lock:
        if dsn in _schema_initialized:
            return
    stmts = [
        """
        CREATE TABLE IF NOT EXISTS check_large_imports (
            id BIGSERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL,
            filename TEXT NOT NULL DEFAULT '',
            sheet_name TEXT NOT NULL DEFAULT '',
            plate_column TEXT NOT NULL DEFAULT '',
            column_order JSONB NOT NULL DEFAULT '[]'::jsonb,
            created_at TIMESTAMPTZ NOT NULL DEFAULT now()
        )
        """,
    ]
    with psycopg.connect(dsn, autocommit=True) as conn:
        for sql in stmts:
            conn.execute(sql)
        _apply_migrations(conn)
        conn.execute("ALTER TABLE check_large_rows ENABLE ROW LEVEL SECURITY")
        conn.execute("ALTER TABLE check_large_rows FORCE ROW LEVEL SECURITY")
        conn.execute("ALTER TABLE check_large_imports ENABLE ROW LEVEL SECURITY")
        conn.execute("ALTER TABLE check_large_imports FORCE ROW LEVEL SECURITY")
        _ensure_peer_group_mirror_and_rls(conn)
        _compact_redundant_gps_in_row_data(conn)
    with _schema_lock:
        _schema_initialized.add(dsn)
    logger.info("check_postgres schema ensured (multi-import)")


def count_user_large_rows_sync(dsn: str, user_id: int, is_admin: bool) -> int:
    ensure_check_pg_schema(dsn)
    with check_pg_tx(dsn, user_id, is_admin) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT COUNT(*)::bigint FROM check_large_rows WHERE user_id = %s",
                (user_id,),
            )
            row = cur.fetchone()
            return int(row[0] if row else 0)


def count_rows_for_user_ids_sync(
    dsn: str, user_id: int, is_admin: bool, target_user_ids: list[int]
) -> int:
    ensure_check_pg_schema(dsn)
    ids = [int(x) for x in (target_user_ids or []) if int(x) > 0]
    if not ids:
        return 0
    with check_pg_tx(dsn, user_id, is_admin) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT COUNT(*)::bigint FROM check_large_rows WHERE user_id = ANY(%s)",
                (ids,),
            )
            row = cur.fetchone()
            return int(row[0] if row else 0)


def list_imports_sync(
    dsn: str,
    user_id: int,
    is_admin: bool,
    *,
    owner_user_id: int | None = None,
) -> list[dict[str, Any]]:
    """قائمة استيرادات المستخدم `owner_user_id` (افتراضياً نفس المُتصل)."""
    ensure_check_pg_schema(dsn)
    owner = int(owner_user_id) if owner_user_id is not None else int(user_id)
    with check_pg_tx(dsn, user_id, is_admin) as conn:
        with conn.cursor(row_factory=dict_row) as cur:
            cur.execute(
                """
                SELECT i.id, i.filename, i.sheet_name, i.plate_column, i.created_at,
                       (SELECT COUNT(*)::int FROM check_large_rows r WHERE r.import_id = i.id) AS row_count
                FROM check_large_imports i
                WHERE i.user_id = %s
                ORDER BY i.created_at ASC, i.id ASC
                """,
                (owner,),
            )
            out = []
            for r in cur.fetchall():
                out.append(
                    {
                        "id": int(r["id"]),
                        "filename": r.get("filename") or "",
                        "sheet_name": r.get("sheet_name") or "",
                        "plate_column": r.get("plate_column") or "",
                        "row_count": int(r["row_count"] or 0),
                        "created_at": r["created_at"].isoformat() if r.get("created_at") else None,
                    }
                )
            return out


def delete_import_sync(dsn: str, user_id: int, is_admin: bool, import_id: int) -> bool:
    ensure_check_pg_schema(dsn)
    with check_pg_tx(dsn, user_id, is_admin) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM check_large_imports WHERE id = %s AND user_id = %s",
                (import_id, user_id),
            )
            return cur.rowcount > 0


def get_stored_large_meta_sync(dsn: str, user_id: int, is_admin: bool) -> dict[str, Any] | None:
    """اتحاد أعمدة كل الاستيرادات + إجمالي الصفوف."""
    ensure_check_pg_schema(dsn)
    imports = list_imports_sync(dsn, user_id, is_admin)
    if not imports:
        return None
    total_rows = sum(i["row_count"] for i in imports)
    if total_rows == 0:
        return None
    seen: set[str] = set()
    union_headers: list[str] = []
    with check_pg_tx(dsn, user_id, is_admin) as conn:
        with conn.cursor(row_factory=dict_row) as cur:
            cur.execute(
                """
                SELECT column_order FROM check_large_imports
                WHERE user_id = %s
                ORDER BY created_at ASC, id ASC
                """,
                (user_id,),
            )
            for r in cur.fetchall():
                co = r["column_order"]
                arr = co if isinstance(co, list) else json.loads(co) if co else []
                for h in arr:
                    hs = str(h).strip() if h else ""
                    if hs and hs not in seen:
                        seen.add(hs)
                        union_headers.append(hs)
    return {
        "headers": union_headers,
        "sheet_name": "",
        "row_count": total_rows,
        "updated_at": imports[-1].get("created_at") if imports else None,
        "imports": imports,
        "has_data": True,
    }


def peer_user_ids_for_check_sync(dsn: str, user_id: int, is_admin: bool) -> list[int]:
    """مستخدم فردي: [نفسه]. عضو مجموعة: كل user_id في نفس المجموعة (مرآة Postgres)."""
    ensure_check_pg_schema(dsn)
    with check_pg_tx(dsn, user_id, is_admin) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT group_id FROM check_mirror_user_groups WHERE user_id = %s",
                (user_id,),
            )
            row = cur.fetchone()
            if row is None:
                return [user_id]
            gid = row[0]
            cur.execute(
                """
                SELECT user_id FROM check_mirror_user_groups
                WHERE group_id = %s
                ORDER BY user_id ASC
                """,
                (gid,),
            )
            peers = [int(r[0]) for r in cur.fetchall()]
            return peers if peers else [user_id]


def get_stored_large_meta_for_check_sync(
    dsn: str, user_id: int, is_admin: bool
) -> dict[str, Any] | None:
    """اتحاد أعمدة كل ملفات المجموعة عند الفرز؛ قائمة الاستيراد في الواجهة تبقى للمستخدم فقط."""
    peers = peer_user_ids_for_check_sync(dsn, user_id, is_admin)
    if len(peers) == 1:
        return get_stored_large_meta_sync(dsn, user_id, is_admin)
    ensure_check_pg_schema(dsn)
    imports_all: list[dict[str, Any]] = []
    for uid in peers:
        imports_all.extend(
            list_imports_sync(dsn, user_id, is_admin, owner_user_id=uid)
        )
    imports_all.sort(key=lambda x: (x.get("created_at") or "", x["id"]))
    if not imports_all:
        return None
    total_rows = sum(i["row_count"] for i in imports_all)
    if total_rows == 0:
        return None
    seen: set[str] = set()
    union_headers: list[str] = []
    with check_pg_tx(dsn, user_id, is_admin) as conn:
        with conn.cursor(row_factory=dict_row) as cur:
            cur.execute(
                """
                SELECT column_order FROM check_large_imports
                WHERE user_id = ANY(%s)
                ORDER BY user_id ASC, created_at ASC, id ASC
                """,
                (peers,),
            )
            for r in cur.fetchall():
                co = r["column_order"]
                arr = co if isinstance(co, list) else json.loads(co) if co else []
                for h in arr:
                    hs = str(h).strip() if h else ""
                    if hs and hs not in seen:
                        seen.add(hs)
                        union_headers.append(hs)
    return {
        "headers": union_headers,
        "sheet_name": "",
        "row_count": total_rows,
        "updated_at": imports_all[-1].get("created_at") if imports_all else None,
        "imports": imports_all,
        "has_data": True,
        "peer_user_ids": peers,
    }


def admin_list_check_storage_sync(dsn: str, admin_user_id: int, is_admin: bool) -> list[dict[str, Any]]:
    if not is_admin:
        return []
    ensure_check_pg_schema(dsn)
    with check_pg_tx(dsn, admin_user_id, True) as conn:
        with conn.cursor(row_factory=dict_row) as cur:
            cur.execute(
                """
                SELECT i.user_id AS user_id,
                       COUNT(DISTINCT i.id)::int AS import_count,
                       COALESCE(
                           (SELECT COUNT(*)::bigint FROM check_large_rows r WHERE r.user_id = i.user_id),
                           0
                       ) AS row_count,
                       MAX(i.created_at) AS last_import
                FROM check_large_imports i
                GROUP BY i.user_id
                ORDER BY last_import DESC NULLS LAST
                """
            )
            out = []
            for r in cur.fetchall():
                out.append(
                    {
                        "user_id": r["user_id"],
                        "import_count": int(r["import_count"] or 0),
                        "row_count": int(r["row_count"] or 0),
                        "updated_at": r["last_import"].isoformat() if r.get("last_import") else None,
                    }
                )
            return out


def _resolve_large_workbook_sheet(
    large_wb: openpyxl.Workbook, large_sheet: str
) -> tuple[openpyxl.worksheet.worksheet.Worksheet, str | None]:
    """
    Choose which worksheet to import from the large file.
    If the client passes a valid sheet name, use it.
    If the workbook has multiple sheets and no valid name was given, use the first sheet only.
    If there is a single sheet, use find_best_sheet (plate column heuristic).
    Returns (worksheet, optional Arabic note for the API/UI).
    """
    names = list(large_wb.sheetnames)
    if not names:
        raise ValueError("الملف الكبير لا يحتوي ورق عمل")
    ls = (large_sheet or "").strip()
    if ls and ls in names:
        return large_wb[ls], None
    if len(names) > 1:
        first = names[0]
        note = (
            f"الملف يحتوي {len(names)} ورقة؛ تم استيراد الورقة الأولى فقط («{first}»)."
        )
        return large_wb[first], note
    return find_best_sheet(large_wb), None


def import_large_workbook_sync(
    dsn: str,
    user_id: int,
    is_admin: bool,
    lc_bytes: bytes,
    password: str,
    large_col: str,
    large_sheet: str,
    source_filename: str,
    group_max_rows_limit: int | None = None,
    user_max_rows_limit: int | None = None,
) -> dict[str, Any]:
    ensure_check_pg_schema(dsn)
    if len(lc_bytes) > CHECK_PG_MAX_LARGE_BYTES:
        raise ValueError("الملف الكبير يتجاوز 15 ميجابايت")

    fname = _safe_filename(source_filename)
    large_wb = load_workbook_maybe_encrypted(lc_bytes, password)
    try:
        large_ws, sheet_selection_note = _resolve_large_workbook_sheet(
            large_wb, large_sheet
        )
        lrows = list(large_ws.iter_rows(values_only=True))
        if not lrows:
            raise ValueError("الملف الكبير فارغ")
        header_l = lrows[0]
        lh = [str(h).strip() if h is not None else "" for h in header_l]
        lc = (large_col or "").strip() or (auto_detect_plate_col(lh) or "")
        if not lc or lc not in lh:
            raise ValueError(
                f"لم يُعثر على عمود اللوحة في الملف الكبير. الأعمدة: {lh}"
            )
        lci = lh.index(lc)
        new_row_count = 0
        for row in lrows[1:]:
            if all(v is None for v in row):
                continue
            rp = row[lci] if lci < len(row) else None
            if normalize_plate(rp):
                new_row_count += 1

        if group_max_rows_limit is not None:
            peers = peer_user_ids_for_check_sync(dsn, user_id, is_admin)
            current_total = count_rows_for_user_ids_sync(
                dsn, user_id, is_admin, peers
            )
            if current_total + new_row_count > int(group_max_rows_limit):
                raise ValueError(
                    "تجاوزت الحد المسموح لعدد الصفوف المخزنة. راجع الحد المتاح لك مع الأدمن."
                )
        else:
            current_total = count_user_large_rows_sync(dsn, user_id, is_admin)
            limit = (
                int(user_max_rows_limit)
                if user_max_rows_limit is not None and int(user_max_rows_limit) > 0
                else CHECK_PG_MAX_ROWS_PER_USER
            )
            if current_total + new_row_count > limit:
                raise ValueError(
                    "تجاوزت الحد المسموح لعدد الصفوف المخزنة. راجع الحد المتاح لك مع الأدمن."
                )

        batch: list[tuple[int, int, str, Json, str]] = []
        batch_size = 2000
        total = 0
        import_id: int | None = None
        with check_pg_tx(dsn, user_id, is_admin) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT 1
                    FROM check_large_imports
                    WHERE user_id = %s
                      AND lower(trim(filename)) = lower(trim(%s))
                    LIMIT 1
                    """,
                    (user_id, fname),
                )
                if cur.fetchone():
                    raise ValueError(
                        f"يوجد ملف بنفس الاسم مسبقاً: {fname}. غيّر الاسم أو احذف الاستيراد القديم أولاً."
                    )
                cur.execute(
                    """
                    INSERT INTO check_large_imports (user_id, filename, sheet_name, plate_column, column_order)
                    VALUES (%s, %s, %s, %s, %s::jsonb)
                    RETURNING id
                    """,
                    (
                        user_id,
                        fname,
                        large_ws.title,
                        lc,
                        json.dumps(lh, ensure_ascii=False),
                    ),
                )
                import_id = int(cur.fetchone()[0])
                for row in lrows[1:]:
                    if all(v is None for v in row):
                        continue
                    rp = row[lci] if lci < len(row) else None
                    pn = normalize_plate(rp)
                    if not pn:
                        continue
                    rd: dict[str, Any] = {}
                    gps_val = ""
                    for hi, h in enumerate(lh):
                        if not h:
                            continue
                        cell = row[hi] if hi < len(row) else None
                        if _norm_header_sim(h) == _norm_header_sim(GPS_HEADER):
                            if not gps_val:
                                gps_val = str(cell).strip() if cell is not None else ""
                            # Keep GPS in dedicated column only to reduce JSONB size.
                            continue
                        rd[h] = _jsonable_cell(cell)
                    batch.append((user_id, import_id, pn, Json(rd), gps_val))
                    total += 1
                    if len(batch) >= batch_size:
                        cur.executemany(
                            """
                            INSERT INTO check_large_rows (user_id, import_id, plate_normalized, row_data, gps)
                            VALUES (%s, %s, %s, %s::jsonb, %s)
                            """,
                            batch,
                        )
                        batch.clear()
                if batch:
                    cur.executemany(
                        """
                        INSERT INTO check_large_rows (user_id, import_id, plate_normalized, row_data, gps)
                        VALUES (%s, %s, %s, %s::jsonb, %s)
                        """,
                        batch,
                    )
        if total == 0 and import_id is not None:
            with check_pg_tx(dsn, user_id, is_admin) as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "DELETE FROM check_large_imports WHERE id = %s AND user_id = %s",
                        (import_id, user_id),
                    )
            raise ValueError("لا توجد صفوف بلوحات صالحة في الملف الكبير")
        out: dict[str, Any] = {
            "row_count": total,
            "headers": lh,
            "sheet_name": large_ws.title,
            "large_col_used": lc,
            "import_id": import_id,
            "filename": fname,
        }
        if sheet_selection_note:
            out["sheet_selection_note"] = sheet_selection_note
        return out
    finally:
        try:
            large_wb.close()
        except Exception:
            pass


def _cell_display(v: Any) -> Any:
    if v is None:
        return None
    return v


def _parse_row_data_dict(rd: Any) -> dict[str, Any]:
    """Normalize JSONB row_data to a str-keyed dict for reliable column lookups."""
    if rd is None:
        return {}
    if isinstance(rd, dict):
        return {str(k) if k is not None else "": v for k, v in rd.items()}
    if isinstance(rd, str):
        try:
            o = json.loads(rd)
            return _parse_row_data_dict(o)
        except Exception:
            return {}
    return {}


def _merge_sql_gps_into_row(row_dict: dict[str, Any], gps_sql: Any) -> dict[str, Any]:
    """
    check_large_rows stores GPS redundantly in row_data and in column `gps`.
    If row_data lost/emptied GPS (legacy rows, drivers), copy from `gps`.
    """
    d = dict(row_dict)
    g_sql = str(gps_sql or "").strip()
    if not g_sql:
        return d
    idx = _row_norm_key_index(d)
    gps_n = _norm_header_sim(GPS_HEADER)
    gps_key = idx.get(gps_n)
    cur = d.get(GPS_HEADER)
    if cur is None and gps_key:
        cur = d.get(gps_key)
    cur_s = "" if cur is None else str(cur).strip()
    if not cur_s:
        if gps_key:
            d[gps_key] = g_sql
        else:
            d[GPS_HEADER] = g_sql
    return d


def _large_row_get(
    row_dict: dict[str, Any],
    col_name: str,
    norm_index: dict[str, str] | None = None,
) -> Any:
    """Lookup cell by export / column_order header; match JSONB keys with Arabic normalization."""
    if not col_name:
        return None
    cn = str(col_name)
    if cn in row_dict:
        return row_dict[cn]
    want = cn.strip().lower()
    for k, v in row_dict.items():
        if str(k).strip().lower() == want:
            return v
    nk = _norm_header_sim(cn)
    if not nk:
        return None
    if norm_index is None:
        norm_index = _row_norm_key_index(row_dict)
    orig = norm_index.get(nk)
    if orig is not None:
        return row_dict.get(orig)
    return None


def _fetch_matches_by_plates_batch(
    conn, distinct_plates: list[str]
) -> dict[str, dict[int, list[dict[str, Any]]]]:
    """
    Load matching large rows for many normalized plates in few queries.
    `distinct_plates` should be unique norms (caller dedupes) to keep payloads small.
    Returns: plate_normalized -> import_id -> row dicts (same shape as before).
    """
    result: dict[str, dict[int, list[dict[str, Any]]]] = {}
    if not distinct_plates:
        return result
    with conn.cursor(row_factory=dict_row) as cur:
        for i in range(0, len(distinct_plates), _CHECK_PLATE_LOOKUP_BATCH):
            chunk = distinct_plates[i : i + _CHECK_PLATE_LOOKUP_BATCH]
            if not chunk:
                continue
            cur.execute(
                """
                SELECT import_id, plate_normalized, row_data, gps
                FROM check_large_rows
                WHERE plate_normalized = ANY(%s) AND import_id IS NOT NULL
                ORDER BY plate_normalized ASC, import_id ASC, id ASC
                """,
                (chunk,),
            )
            for row in cur.fetchall():
                pn = str(row.get("plate_normalized") or "").strip()
                if not pn:
                    continue
                iid = row.get("import_id")
                if iid is None:
                    continue
                iid_i = int(iid)
                d = _merge_sql_gps_into_row(
                    _parse_row_data_dict(row.get("row_data")),
                    row.get("gps"),
                )
                bucket = result.setdefault(pn, {})
                bucket.setdefault(iid_i, []).append(d)
    return result


def _load_imports_ordered_for_peers(
    conn, peer_user_ids: list[int]
) -> list[dict[str, Any]]:
    if not peer_user_ids:
        return []
    with conn.cursor(row_factory=dict_row) as cur:
        cur.execute(
            """
            SELECT id, filename, sheet_name, plate_column, column_order
            FROM check_large_imports
            WHERE user_id = ANY(%s)
            ORDER BY user_id ASC, created_at ASC, id ASC
            """,
            (peer_user_ids,),
        )
        rows = []
        for r in cur.fetchall():
            co = r["column_order"]
            if isinstance(co, str):
                co = json.loads(co)
            rows.append(
                {
                    "id": int(r["id"]),
                    "filename": r.get("filename") or "",
                    "sheet_name": r.get("sheet_name") or "",
                    "plate_column": r.get("plate_column") or "",
                    "column_order": list(co) if co else [],
                }
            )
        return rows


def _ws_set_col_widths(ws, max_col: int, width: float = 18.0) -> None:
    for c in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = width


def _plate_col_matches(col_name: str, plate_column: str) -> bool:
    if not plate_column or not col_name:
        return False
    return _norm_header_sim(col_name) == _norm_header_sim(plate_column)


def _write_row(
    ws,
    row_idx: int,
    values: list[Any],
    *,
    header: bool,
    header_fills: list[str] | None,
    body_fill_large: PatternFill,
    body_fill_small: PatternFill,
    body_fill_plate: PatternFill,
    fonts_h: Font,
    fonts_b: Font,
    align_h: Alignment,
    align_b: Alignment,
    border: Border,
    col_sources: list[str],
) -> int:
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = border
        src = col_sources[col_idx - 1] if col_idx - 1 < len(col_sources) else "large"
        if header:
            cell.font = fonts_h
            cell.alignment = align_h
            if header_fills and col_idx - 1 < len(header_fills):
                cell.fill = PatternFill("solid", start_color=header_fills[col_idx - 1])
            else:
                cell.fill = (
                    PatternFill("solid", start_color="1E40AF")
                    if src == "large"
                    else PatternFill("solid", start_color="166534")
                )
        else:
            cell.font = fonts_b
            cell.alignment = align_b
            if src == "plate":
                cell.fill = body_fill_plate
            elif src == "small":
                cell.fill = body_fill_small
            else:
                cell.fill = body_fill_large
    return row_idx + 1


def run_check_plates_postgres_sync(
    dsn: str,
    user_id: int,
    is_admin: bool,
    sc_bytes: bytes,
    password: str,
    small_col: str,
    small_sheet: str,
    large_export_cols: list[str] | None,
    small_export_cols: list[str] | None,
) -> dict:
    ensure_check_pg_schema(dsn)
    peer_ids = peer_user_ids_for_check_sync(dsn, user_id, is_admin)
    meta = get_stored_large_meta_for_check_sync(dsn, user_id, is_admin)
    if not meta or meta.get("row_count", 0) == 0:
        return {
            "kind": "json",
            "status_code": 400,
            "body": {
                "detail": "لا توجد بيانات ملفات كبيرة مخزّنة. استورد ملفاً كبيراً أولاً.",
                "code": "NO_STORED_LARGE",
            },
        }

    union_headers: list[str] = list(meta.get("headers") or [])

    small_wb = None
    try:
        small_wb = openpyxl.load_workbook(
            io.BytesIO(sc_bytes), read_only=True, data_only=True
        )
        small_ws = (
            small_wb[small_sheet]
            if small_sheet and small_sheet in small_wb.sheetnames
            else find_best_sheet(small_wb)
        )
        srows = small_ws.iter_rows(values_only=True)
        header_s = next(srows, None)
        if header_s is None:
            raise ValueError("الملف الصغير فارغ")
        sh = [str(h).strip() if h is not None else "" for h in header_s]
        row2 = next(srows, None)
        row3 = next(srows, None)
        detected_small = (auto_detect_plate_col(sh) or auto_detect_plate_col_from_row3(sh, row3) or "")
        sc = small_col.strip() or detected_small
        if not sc or sc not in sh:
            return {
                "kind": "json",
                "status_code": 422,
                "body": {
                    "detail": f"لم يُعثر على عمود اللوحة في الملف الصغير. الأعمدة: {sh}",
                    "headers": sh,
                    "code": "COL_NOT_FOUND_SMALL",
                },
            }
        se_cols = _norm_small_export_cols(list(small_export_cols or []), sh)
        se_idx = [sh.index(c) for c in se_cols]
        sci = sh.index(sc)

        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        body_font = Font(name="Arial", size=10, color="000000")
        body_fill_large = PatternFill("solid", start_color="DBEAFE")
        body_fill_small = PatternFill("solid", start_color="DCFCE7")
        body_fill_plate = PatternFill("solid", start_color="E0E7FF")
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_cell = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = _thin_border()

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "التطابقات"
        ws.sheet_view.rightToLeft = True
        row = 1

        preview_rows: list[list[str]] = []
        preview_sections: list[dict[str, Any]] = []
        matched_rows_count = 0
        matched_plate_hits = 0
        unmatched_plates = 0
        section_started: dict[int, bool] = {}
        multi_file_notes: list[tuple[str, str, list[Any], dict[int, list[dict]]]] = []

        with check_pg_tx(dsn, user_id, is_admin) as conn:
            imports = _load_imports_ordered_for_peers(conn, peer_ids)
            if not imports:
                return {
                    "kind": "json",
                    "status_code": 400,
                    "body": {"detail": "لا توجد استيرادات.", "code": "NO_IMPORTS"},
                }
            imp_by_id = {i["id"]: i for i in imports}

            # Pass 1a: read small file rows (order preserved); Pass 1b: one batched DB lookup per plate chunk.
            pending: list[dict[str, Any]] = []
            for srow in itertools.chain((r for r in (row2, row3) if r is not None), srows):
                if all(v is None for v in srow):
                    continue
                rp = srow[sci] if sci < len(srow) else None
                norm = normalize_plate(rp)
                if not norm:
                    continue
                small_vals = [(srow[i] if i < len(srow) else None) for i in se_idx]
                plate_disp = str(rp).strip() if rp is not None else norm
                pending.append(
                    {
                        "norm": norm,
                        "plate_disp": plate_disp,
                        "small_vals": small_vals,
                    }
                )
            unique_norms = list(dict.fromkeys([p["norm"] for p in pending]))
            norm_cache = _fetch_matches_by_plates_batch(conn, unique_norms)

            small_entries: list[dict[str, Any]] = []
            for p in pending:
                norm = p["norm"]
                groups = norm_cache.get(norm) or {}
                if not groups:
                    unmatched_plates += 1
                    continue
                matched_plate_hits += 1
                plate_disp = p["plate_disp"]
                small_vals = p["small_vals"]
                if len(groups) > 1:
                    multi_file_notes.append((norm, plate_disp, small_vals, dict(groups)))
                small_entries.append(
                    {
                        "norm": norm,
                        "plate_disp": plate_disp,
                        "small_vals": small_vals,
                        "groups": groups,
                    }
                )

            # Pass 2: for each large-file import in order, emit all matching rows for that
            # file only — avoids interleaving sections so rows stay under the correct file.
            for imp in imports:
                iid = imp["id"]
                lh = [
                    str(h).strip()
                    for h in (imp.get("column_order") or [])
                    if h is not None and str(h).strip()
                ]
                if not lh:
                    lh = list(union_headers)
                sec_le = _map_export_headers_to_sheet(
                    list(large_export_cols or []), lh
                )
                if not sec_le:
                    sec_le = _norm_large_export_cols(
                        list(large_export_cols or []), union_headers
                    )
                if not sec_le:
                    sec_le = list(union_headers)
                pc_l = (imp.get("plate_column") or "").strip()
                for ent in small_entries:
                    lst = ent["groups"].get(iid) or []
                    if not lst:
                        continue
                    small_vals = ent["small_vals"]
                    if not section_started.get(iid):
                        color_hex = FILE_SECTION_HEADER_FILLS[
                            imports.index(imp) % len(FILE_SECTION_HEADER_FILLS)
                        ]
                        title_fill = PatternFill("solid", start_color=color_hex)
                        sn = (imp.get("sheet_name") or "").strip()
                        title_txt = f"📁 {imp['filename']}"
                        if sn:
                            title_txt = f"{title_txt} — ورقة: {sn}"
                        cell = ws.cell(row=row, column=1, value=title_txt)
                        cell.fill = title_fill
                        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
                        cell.border = border
                        row += 1
                        hdr_vals = [
                            _strip_small_word_from_header_title(c) for c in sec_le
                        ] + [
                            _strip_small_word_from_header_title(c) for c in se_cols
                        ]
                        hdr_src = ["large"] * len(sec_le) + ["small"] * len(se_cols)
                        hdr_colors = [color_hex] * len(hdr_vals)
                        row = _write_row(
                            ws,
                            row,
                            hdr_vals,
                            header=True,
                            header_fills=hdr_colors,
                            body_fill_large=body_fill_large,
                            body_fill_small=body_fill_small,
                            body_fill_plate=body_fill_plate,
                            fonts_h=header_font,
                            fonts_b=body_font,
                            align_h=align_header,
                            align_b=align_cell,
                            border=border,
                            col_sources=hdr_src,
                        )
                        section_started[iid] = True
                        title_ui = (imp.get("filename") or "").strip() or "ملف"
                        if sn:
                            title_ui = f"{title_ui} — ورقة: {sn}"
                        plate_idx_sec: list[int] = []
                        for i, c in enumerate(sec_le):
                            if _plate_col_matches(c, pc_l):
                                plate_idx_sec.append(i)
                        for i, c in enumerate(se_cols):
                            if _plate_col_matches(c, sc):
                                plate_idx_sec.append(len(sec_le) + i)
                        preview_sections.append(
                            {
                                "title": title_ui,
                                "headers": [str(h) if h is not None else "" for h in hdr_vals],
                                "col_sources": list(hdr_src),
                                "plate_column_indices": plate_idx_sec,
                                "rows": [],
                            }
                        )

                    for ld in lst:
                        nidx = _row_norm_key_index(ld)
                        large_vals = [
                            _cell_display(_large_row_get(ld, c, nidx))
                            for c in sec_le
                        ]
                        row_vals = large_vals + small_vals
                        row_src = ["large"] * len(sec_le) + ["small"] * len(se_cols)
                        row = _write_row(
                            ws,
                            row,
                            row_vals,
                            header=False,
                            header_fills=None,
                            body_fill_large=body_fill_large,
                            body_fill_small=body_fill_small,
                            body_fill_plate=body_fill_plate,
                            fonts_h=header_font,
                            fonts_b=body_font,
                            align_h=align_header,
                            align_b=align_cell,
                            border=border,
                            col_sources=row_src,
                        )
                        matched_rows_count += 1
                        if len(preview_rows) < PREVIEW_MAX_ROWS:
                            pr = [
                                "" if v is None else str(v).strip() for v in row_vals
                            ]
                            preview_rows.append(pr)
                            if preview_sections:
                                preview_sections[-1]["rows"].append(pr)

            if multi_file_notes:
                row += 1
                ws.cell(row=row, column=1, value="⚠ لوحات ظهرت في أكثر من ملف كبير")
                ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="9A3412")
                row += 1
                sub_hdr = ["رقم اللوحة", "الملفات", "ملاحظة"]
                row = _write_row(
                    ws,
                    row,
                    sub_hdr,
                    header=True,
                    header_fills=["B91C1C", "B91C1C", "B91C1C"],
                    body_fill_large=body_fill_large,
                    body_fill_small=body_fill_small,
                    body_fill_plate=body_fill_plate,
                    fonts_h=header_font,
                    fonts_b=body_font,
                    align_h=align_header,
                    align_b=align_cell,
                    border=border,
                    col_sources=["plate", "large", "large"],
                )
                seen_norm: set[str] = set()
                for norm, plate_disp, smv, grp in multi_file_notes:
                    if norm in seen_norm:
                        continue
                    seen_norm.add(norm)
                    names = []
                    for iid in sorted(grp.keys()):
                        im = imp_by_id.get(iid)
                        if im:
                            names.append(im["filename"])
                    note = f"تطابق في {len(grp)} ملف(ات)"
                    row = _write_row(
                        ws,
                        row,
                        [plate_disp, "; ".join(names), note],
                        header=False,
                        header_fills=None,
                        body_fill_large=body_fill_large,
                        body_fill_small=body_fill_small,
                        body_fill_plate=body_fill_plate,
                        fonts_h=header_font,
                        fonts_b=body_font,
                        align_h=align_header,
                        align_b=align_cell,
                        border=border,
                        col_sources=["plate", "large", "large"],
                    )

        max_col = 1
        if ws.max_column:
            max_col = ws.max_column
        _ws_set_col_widths(ws, max_col, 20.0)

        if not matched_rows_count:
            return {
                "kind": "json",
                "status_code": 200,
                "body": {
                    "detail": "لا توجد تطابقات",
                    "matched": 0,
                    "unmatched": unmatched_plates,
                    "large_col_used": "(مخزّن — كل الملفات)",
                    "small_col_used": sc,
                },
            }

        le_preview = _norm_large_export_cols(list(large_export_cols or []), union_headers)
        display_headers = [
            _strip_small_word_from_header_title(c) for c in le_preview
        ] + [_strip_small_word_from_header_title(c) for c in se_cols]
        col_sources_preview = ["large"] * len(le_preview) + ["small"] * len(se_cols)
        plate_indices_flat: list[int] = []
        if imports:
            pcl0 = (imports[0].get("plate_column") or "").strip()
            for i, c in enumerate(le_preview):
                if _plate_col_matches(c, pcl0):
                    plate_indices_flat.append(i)
            for i, c in enumerate(se_cols):
                if _plate_col_matches(c, sc):
                    plate_indices_flat.append(len(le_preview) + i)

        content = workbook_to_bytes(wb)
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"التطابقات_{ts}.xlsx"
        return {
            "kind": "xlsx",
            "content": content,
            "filename": filename,
            "preview": {
                "headers": display_headers,
                "col_sources": col_sources_preview,
                "plate_column_indices": plate_indices_flat,
                "rows": preview_rows,
                "sections": preview_sections,
                "total_rows": matched_rows_count,
                "truncated": matched_rows_count > PREVIEW_MAX_ROWS,
                "stats": {
                    "matched_rows": matched_rows_count,
                    "matched_plate_hits": matched_plate_hits,
                    "unmatched_plates": unmatched_plates,
                    "multi_file_plates": len({m[0] for m in multi_file_notes}),
                },
            },
        }
    finally:
        if small_wb:
            try:
                small_wb.close()
            except Exception:
                pass


def collect_gps_vehicles_stored_sync(
    dsn: str,
    user_id: int,
    is_admin: bool,
    sc_bytes: bytes,
    small_col: str,
    small_sheet: str,
) -> dict[str, Any]:
    ensure_check_pg_schema(dsn)
    peer_ids = peer_user_ids_for_check_sync(dsn, user_id, is_admin)
    meta = get_stored_large_meta_for_check_sync(dsn, user_id, is_admin)
    if not meta or not meta.get("headers"):
        return {"detail": "لا توجد بيانات ملف كبير مخزّنة", "vehicles": []}

    union_h: list[str] = list(meta["headers"])
    gps_col = GPS_HEADER if GPS_HEADER in union_h else None
    if not gps_col:
        return {"detail": "لا يوجد عمود GPS في البيانات المخزّنة", "vehicles": []}

    date_col = next((h for h in union_h if "تاريخ" in h), None)
    type_col = next((h for h in union_h if "نوع" in h), None)
    notes_col = next((h for h in union_h if "ملاحظات" in h), None)

    small_wb = openpyxl.load_workbook(io.BytesIO(sc_bytes), read_only=True, data_only=True)
    try:
        small_ws = (
            small_wb[small_sheet]
            if small_sheet and small_sheet in small_wb.sheetnames
            else find_best_sheet(small_wb)
        )
        sd = list(small_ws.iter_rows(values_only=True))
        if not sd:
            return {"detail": "الملف الصغير فارغ", "vehicles": []}
        sh = [str(h).strip() if h is not None else "" for h in sd[0]]
        row3 = sd[2] if len(sd) > 2 else None
        sc = small_col.strip() or (auto_detect_plate_col(sh) or auto_detect_plate_col_from_row3(sh, row3) or "")
        if not sc or sc not in sh:
            return {
                "detail": f"لم يُعثر على عمود اللوحة في الملف الصغير. الأعمدة: {sh}",
                "vehicles": [],
            }
        sci = sh.index(sc)
        matched: list[dict[str, Any]] = []
        seen: set[tuple[str, str, str]] = set()

        gps_rows: list[tuple[Any, str]] = []
        for row in sd[1:]:
            if all(v is None for v in row):
                continue
            rp = row[sci] if sci < len(row) else None
            norm = normalize_plate(rp)
            if not norm:
                continue
            gps_rows.append((rp, norm))

        with check_pg_tx(dsn, user_id, is_admin) as conn:
            imports = _load_imports_ordered_for_peers(conn, peer_ids)
            imp_name = {i["id"]: i["filename"] for i in imports}
            unique_gps = list(dict.fromkeys([t[1] for t in gps_rows]))
            norm_cache = _fetch_matches_by_plates_batch(conn, unique_gps)
            for rp, norm in gps_rows:
                groups = norm_cache.get(norm) or {}
                for iid, rds in groups.items():
                    fn = imp_name.get(iid, "")
                    for ld in rds:
                        nix = _row_norm_key_index(ld)
                        gps = str(
                            _large_row_get(ld, gps_col or "", nix) or ""
                        ).strip()
                        plate = str(rp or "").strip()
                        key = (plate, gps, fn)
                        if key in seen:
                            continue
                        seen.add(key)
                        matched.append(
                            {
                                "plate": plate,
                                "gps": gps,
                                "date": str(
                                    _large_row_get(ld, date_col or "", nix) or ""
                                ).strip()
                                if date_col
                                else "",
                                "vehicle_type": str(
                                    _large_row_get(ld, type_col or "", nix) or ""
                                ).strip()
                                if type_col
                                else "",
                                "notes": (
                                    str(
                                        _large_row_get(
                                            ld, notes_col or "", nix
                                        )
                                        or ""
                                    ).strip()
                                    if notes_col
                                    else ""
                                )
                                + ((" " + fn) if fn else ""),
                            }
                        )

        valid = []
        for item in matched:
            gps = item.get("gps", "")
            if not gps or gps == "None" or "," not in gps:
                continue
            parts = gps.split(",")
            try:
                float(parts[0].strip())
                float(parts[1].strip())
                valid.append(item)
            except Exception:
                continue

        if not valid:
            return {"detail": "لا توجد تطابقات", "vehicles": [], "total": 0, "skipped": 0}

        return {
            "vehicles": valid,
            "total": len(valid),
            "skipped": len(matched) - len(valid),
        }
    finally:
        try:
            small_wb.close()
        except Exception:
            pass


# ── Admin: browse / export check_large_* (RLS bypass via is_admin) ───────────

ADMIN_IMPORT_PREVIEW_MAX = 200
ADMIN_IMPORT_PAGE_MAX = 500


def admin_list_imports_detailed_sync(
    dsn: str, admin_user_id: int, is_admin: bool
) -> list[dict[str, Any]]:
    if not is_admin:
        return []
    ensure_check_pg_schema(dsn)
    with check_pg_tx(dsn, admin_user_id, True) as conn:
        with conn.cursor(row_factory=dict_row) as cur:
            cur.execute(
                """
                SELECT i.id AS import_id,
                       i.user_id AS user_id,
                       i.filename AS filename,
                       i.sheet_name AS sheet_name,
                       i.plate_column AS plate_column,
                       i.created_at AS created_at,
                       COUNT(r.id)::bigint AS row_count
                FROM check_large_imports i
                LEFT JOIN check_large_rows r ON r.import_id = i.id AND r.user_id = i.user_id
                GROUP BY i.id, i.user_id, i.filename, i.sheet_name, i.plate_column, i.created_at
                ORDER BY i.created_at DESC NULLS LAST, i.id DESC
                """
            )
            out: list[dict[str, Any]] = []
            for r in cur.fetchall():
                out.append(
                    {
                        "import_id": int(r["import_id"]),
                        "user_id": int(r["user_id"]),
                        "filename": r.get("filename") or "",
                        "sheet_name": r.get("sheet_name") or "",
                        "plate_column": r.get("plate_column") or "",
                        "row_count": int(r["row_count"] or 0),
                        "created_at": r["created_at"].isoformat()
                        if r.get("created_at")
                        else None,
                    }
                )
            return out


def admin_get_import_meta_sync(
    dsn: str, admin_user_id: int, is_admin: bool, import_id: int
) -> dict[str, Any] | None:
    if not is_admin or import_id <= 0:
        return None
    ensure_check_pg_schema(dsn)
    with check_pg_tx(dsn, admin_user_id, True) as conn:
        with conn.cursor(row_factory=dict_row) as cur:
            cur.execute(
                """
                SELECT id, user_id, filename, sheet_name, plate_column, column_order
                FROM check_large_imports
                WHERE id = %s
                """,
                (import_id,),
            )
            r = cur.fetchone()
            if not r:
                return None
            co = r["column_order"]
            if isinstance(co, str):
                co = json.loads(co)
            cols = list(co) if co else []
            return {
                "import_id": int(r["id"]),
                "user_id": int(r["user_id"]),
                "filename": r.get("filename") or "",
                "sheet_name": r.get("sheet_name") or "",
                "plate_column": r.get("plate_column") or "",
                "column_order": cols,
            }


def admin_list_import_rows_page_sync(
    dsn: str,
    admin_user_id: int,
    is_admin: bool,
    import_id: int,
    offset: int,
    limit: int,
) -> tuple[list[dict[str, Any]], int]:
    if not is_admin or import_id <= 0:
        return [], 0
    limit = max(1, min(int(limit or 20), ADMIN_IMPORT_PAGE_MAX))
    offset = max(0, int(offset or 0))
    ensure_check_pg_schema(dsn)
    with check_pg_tx(dsn, admin_user_id, True) as conn:
        with conn.cursor(row_factory=dict_row) as cur:
            cur.execute(
                "SELECT COUNT(*)::bigint AS cnt FROM check_large_rows WHERE import_id = %s",
                (import_id,),
            )
            total = int(cur.fetchone()["cnt"] or 0)
            cur.execute(
                """
                SELECT id, plate_normalized, row_data, gps
                FROM check_large_rows
                WHERE import_id = %s
                ORDER BY id ASC
                LIMIT %s OFFSET %s
                """,
                (import_id, limit, offset),
            )
            rows: list[dict[str, Any]] = []
            for row in cur.fetchall():
                rd = _merge_sql_gps_into_row(
                    _parse_row_data_dict(row.get("row_data")),
                    row.get("gps"),
                )
                rows.append(
                    {
                        "id": int(row["id"]),
                        "plate_normalized": row.get("plate_normalized") or "",
                        "gps": row.get("gps") or "",
                        "row_data": rd if isinstance(rd, dict) else {},
                    }
                )
            return rows, total


def admin_write_import_csv_tempfile_sync(
    dsn: str, admin_user_id: int, is_admin: bool, import_id: int
) -> tuple[str, str]:
    """Write UTF-8 CSV with BOM to a temp path. Returns (path, safe_filename)."""
    if not is_admin or import_id <= 0:
        raise ValueError("invalid export")
    meta = admin_get_import_meta_sync(dsn, admin_user_id, is_admin, import_id)
    if not meta:
        raise ValueError("import not found")
    headers = ["id", "plate_normalized", "gps"] + list(meta.get("column_order") or [])
    base = _safe_filename(meta.get("filename") or f"import_{import_id}")
    if base.lower().endswith(".xlsx"):
        base = base[:-5]
    out_name = f"{base}_user{meta['user_id']}_import{import_id}.csv"

    ensure_check_pg_schema(dsn)
    fd, path = tempfile.mkstemp(suffix=".csv", prefix="check_export_")
    try:
        with os.fdopen(fd, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(headers)
            batch = 8000
            with check_pg_tx(dsn, admin_user_id, True) as conn:
                with conn.cursor(row_factory=dict_row) as cur:
                    cur.execute(
                        """
                        SELECT id, plate_normalized, row_data, gps
                        FROM check_large_rows
                        WHERE import_id = %s
                        ORDER BY id ASC
                        """,
                        (import_id,),
                    )
                    while True:
                        chunk = cur.fetchmany(batch)
                        if not chunk:
                            break
                        for row in chunk:
                            rd = _merge_sql_gps_into_row(
                                _parse_row_data_dict(row.get("row_data")),
                                row.get("gps"),
                            )
                            vals: list[Any] = [
                                int(row["id"]),
                                row.get("plate_normalized") or "",
                                row.get("gps") or "",
                            ]
                            for h in meta.get("column_order") or []:
                                vals.append(_cell_display(rd.get(h)))
                            w.writerow(vals)
        return path, out_name
    except Exception:
        try:
            os.unlink(path)
        except Exception:
            pass
        raise
