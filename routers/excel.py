import asyncio
import io
import json
import logging
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import APIRouter, Depends, Form, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse

from dependencies.auth import get_current_user
from services.plate_utils import normalize_plate_value
from services.excel_utils import apply_excel_style, workbook_to_bytes_async
from services.upload_security import MAX_EXCEL_BYTES, save_upload_to_temp_with_limit
from urllib.parse import quote


logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/api",
    tags=["excel"],
    dependencies=[Depends(get_current_user)],
)

_EXPORT_HEADERS = [
    "رقم اللوحة",
    "GPS",
    "تاريخ التسجيل",
    "الحي",
    "الشارع",
    "ملاحظات",
    "نوع السيارة",
    "اسم المسجّل",
    "موقع الشارع",
]
_COL_WIDTHS = [22, 26, 18, 18, 22, 40, 14, 20, 26]
_EXPORT_HEADERS_SET = {h.strip() for h in _EXPORT_HEADERS}

# جدول جلسة التشيك (صفحة التشيك فقط): بدون نوع السيارة وموقع الشارع وملاحظات
_CHECK_SESSION_EXPORT_HEADERS = [
    "رقم اللوحة",
    "GPS",
    "تاريخ التسجيل",
    "الحي",
    "الشارع",
    "اسم المسجّل",
]
_CHECK_SESSION_COL_WIDTHS = [22, 26, 18, 18, 22, 20]
_CHECK_SESSION_EXPORT_HEADERS_SET = {h.strip() for h in _CHECK_SESSION_EXPORT_HEADERS}


def _clean_sheet_name(name: str) -> str:
    for ch in r'/\?*[]':
        name = name.replace(ch, "")
    return (name or "بيانات المركبات")[:31]

def _content_disposition(filename_utf8: str, fallback_ascii: str = "export.xlsx") -> str:
    """
    Starlette encodes header values as latin-1; non-ASCII must be percent-encoded.
    Provide an ASCII fallback filename for older clients.
    """
    encoded = quote(filename_utf8, safe="")
    return f'attachment; filename="{fallback_ascii}"; filename*=UTF-8\'\'{encoded}'

def _mid_gps_value(valid_rows: list[dict]) -> str:
    """
    Pick a representative GPS value from the middle of the recording.
    Example: if there are 100 cars, pick item #50 (1-based) -> index 49.
    """
    if not valid_rows:
        return ""
    mid_idx = (len(valid_rows) - 1) // 2
    return str(valid_rows[mid_idx].get("gps", "") or "").strip()


def _row_street_location(r: dict) -> str:
    """Per-row موقع الشارع from API; legacy rows fall back to that row's GPS."""
    s = str(r.get("street_location", "") or "").strip()
    if s:
        return s
    return _mid_gps_value([r])


def _open_workbook_readonly(source: bytes | str):
    if isinstance(source, (str, os.PathLike)):
        return openpyxl.load_workbook(str(source), read_only=True, data_only=True)
    return openpyxl.load_workbook(io.BytesIO(source), read_only=True, data_only=True)


def _parse_append_excel_sync(source: bytes | str) -> tuple[list[dict], int]:
    """
    Parse append-only export file with strict headers.
    Accepts only the canonical export headers (no missing/extra columns).
    """
    wb = _open_workbook_readonly(source)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            raise ValueError("الملف فارغ")
        headers = [str(c).strip() if c is not None else "" for c in header_row]
        headers_no_empty = [h for h in headers if h]
        if len(headers_no_empty) != len(_EXPORT_HEADERS_SET) or set(headers_no_empty) != _EXPORT_HEADERS_SET:
            raise ValueError(
                "أعمدة الملف غير مطابقة. المطلوب فقط: "
                + "، ".join(_EXPORT_HEADERS)
            )
        idx = {h: headers.index(h) for h in _EXPORT_HEADERS}
        out: list[dict] = []
        for row in rows_iter:
            if row is None or all(v is None for v in row):
                continue

            def cell(name: str) -> str:
                ci = idx[name]
                if ci >= len(row) or row[ci] is None:
                    return ""
                return str(row[ci]).strip()

            out.append(
                {
                    "full_plate": cell("رقم اللوحة"),
                    "gps": cell("GPS"),
                    "recording_date": cell("تاريخ التسجيل"),
                    "district_name": cell("الحي"),
                    "street_name": cell("الشارع"),
                    "location_details": cell("ملاحظات"),
                    "vehicle_type": cell("نوع السيارة"),
                    "recorder_name": cell("اسم المسجّل"),
                    "street_location": cell("موقع الشارع"),
                }
            )
        return out, len(out)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _parse_check_session_append_sync(source: bytes | str) -> tuple[list[dict], int]:
    """Parse append file for check-session export (6 columns)."""
    wb = _open_workbook_readonly(source)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            raise ValueError("الملف فارغ")
        headers = [str(c).strip() if c is not None else "" for c in header_row]
        headers_no_empty = [h for h in headers if h]
        if (
            len(headers_no_empty) != len(_CHECK_SESSION_EXPORT_HEADERS_SET)
            or set(headers_no_empty) != _CHECK_SESSION_EXPORT_HEADERS_SET
        ):
            raise ValueError(
                "أعمدة الملف غير مطابقة. المطلوب فقط: "
                + "، ".join(_CHECK_SESSION_EXPORT_HEADERS)
            )
        idx = {h: headers.index(h) for h in _CHECK_SESSION_EXPORT_HEADERS}
        out: list[dict] = []
        for row in rows_iter:
            if row is None or all(v is None for v in row):
                continue

            def cell(name: str) -> str:
                ci = idx[name]
                if ci >= len(row) or row[ci] is None:
                    return ""
                return str(row[ci]).strip()

            out.append(
                {
                    "full_plate": cell("رقم اللوحة"),
                    "gps": cell("GPS"),
                    "recording_date": cell("تاريخ التسجيل"),
                    "district_name": cell("الحي"),
                    "street_name": cell("الشارع"),
                    "recorder_name": cell("اسم المسجّل"),
                }
            )
        return out, len(out)
    finally:
        try:
            wb.close()
        except Exception:
            pass


@router.post("/export-excel")
async def export_excel(
    rows_json:  str = Form("[]"),
    sheet_name: str = Form("بيانات المركبات"),
):
    sheet_name = _clean_sheet_name(sheet_name.strip())

    try:
        rows = json.loads(rows_json)
    except Exception:
        raise HTTPException(status_code=400, detail="تنسيق JSON خاطئ")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.rightToLeft = True

    hf    = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    hfill = PatternFill("solid", start_color="1F4E79")
    ha    = Alignment(horizontal="center", vertical="center")
    ca    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    df    = Font(name="Arial", size=11)
    pf    = Font(name="Arial", size=11, bold=True)  # plate font
    thin  = Side(style="thin", color="BFBFBF")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    fe    = PatternFill("solid", start_color="D6E4F0")
    fo    = PatternFill("solid", start_color="FFFFFF")

    for col, h in enumerate(_EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill
        cell.alignment = ha; cell.border = brd
    ws.row_dimensions[1].height = 30

    # Filter invalid plates
    valid_rows = []
    for r in rows:
        normalized, ok = normalize_plate_value(full_raw=r.get("full_plate", ""))
        if not ok:
            continue
        rr = dict(r)
        rr["full_plate"] = normalized
        valid_rows.append(rr)

    for i, r in enumerate(valid_rows, 1):
        fill = fe if i % 2 == 0 else fo
        vals = [
            r.get("full_plate", ""),
            r.get("gps", ""),
            r.get("recording_date", ""),
            r.get("district_name", ""),
            r.get("street_name", "غير محدد"),
            r.get("location_details", ""),
            r.get("vehicle_type", "ملاكى"),
            r.get("recorder_name", ""),
            _row_street_location(r),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i + 1, column=col, value=v)
            cell.font = pf if col == 1 else df
            cell.alignment = ca
            cell.border = brd; cell.fill = fill

    for col, w in zip("ABCDEFGHI", _COL_WIDTHS):
        ws.column_dimensions[col].width = w

    content = await workbook_to_bytes_async(wb)
    filename = f"تفريغ_{sheet_name}.xlsx"

    return StreamingResponse(
        io.BytesIO(content),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": _content_disposition(filename, "tafreegh.xlsx")},
    )


@router.post("/export-check-session")
async def export_check_session(
    rows_json: str = Form("[]"),
    sheet_name: str = Form("بيانات مركبات الجلسة"),
):
    """Excel for check page session table only — 6 columns (no نوع السيارة / موقع الشارع / ملاحظات)."""
    sheet_name = _clean_sheet_name(sheet_name.strip() or "بيانات مركبات الجلسة")

    try:
        rows = json.loads(rows_json)
    except Exception:
        raise HTTPException(status_code=400, detail="تنسيق JSON خاطئ")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.rightToLeft = True

    hf = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    hfill = PatternFill("solid", start_color="1F4E79")
    ha = Alignment(horizontal="center", vertical="center")
    ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
    df = Font(name="Arial", size=11)
    pf = Font(name="Arial", size=11, bold=True)
    thin = Side(style="thin", color="BFBFBF")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    fe = PatternFill("solid", start_color="D6E4F0")
    fo = PatternFill("solid", start_color="FFFFFF")

    for col, h in enumerate(_CHECK_SESSION_EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = ha
        cell.border = brd
    ws.row_dimensions[1].height = 30

    valid_rows: list[dict] = []
    for r in rows:
        normalized, ok = normalize_plate_value(full_raw=r.get("full_plate", ""))
        if not ok:
            continue
        rr = dict(r)
        rr["full_plate"] = normalized
        valid_rows.append(rr)

    for i, r in enumerate(valid_rows, 1):
        fill = fe if i % 2 == 0 else fo
        vals = [
            r.get("full_plate", ""),
            r.get("gps", ""),
            r.get("recording_date", ""),
            r.get("district_name", ""),
            r.get("street_name", "غير محدد"),
            r.get("recorder_name", ""),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i + 1, column=col, value=v)
            cell.font = pf if col == 1 else df
            cell.alignment = ca
            cell.border = brd
            cell.fill = fill

    for col, w in zip("ABCDEF", _CHECK_SESSION_COL_WIDTHS):
        ws.column_dimensions[col].width = w

    content = await workbook_to_bytes_async(wb)
    filename = f"تشيك_جلسة_{sheet_name}.xlsx"

    return StreamingResponse(
        io.BytesIO(content),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": _content_disposition(filename, "check_session.xlsx")},
    )


@router.post("/export-field-check")
async def export_field_check(
    rows_json:  str = Form("[]"),
    sheet_name: str = Form("التشيك الميداني"),
):
    """Same as export-excel but with a different default sheet name."""
    sheet_name = _clean_sheet_name(sheet_name.strip() or "التشيك الميداني")

    try:
        rows = json.loads(rows_json)
    except Exception:
        raise HTTPException(status_code=400, detail="تنسيق JSON خاطئ")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.rightToLeft = True

    hf    = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    hfill = PatternFill("solid", start_color="0D6B5E")   # teal for field-check
    ha    = Alignment(horizontal="center", vertical="center")
    ca    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    df    = Font(name="Arial", size=11)
    pf    = Font(name="Arial", size=11, bold=True)  # plate font
    thin  = Side(style="thin", color="BFBFBF")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    fe    = PatternFill("solid", start_color="E0F2F1")
    fo    = PatternFill("solid", start_color="FFFFFF")

    headers = [
        "رقم اللوحة",
        "GPS",
        "تاريخ التسجيل",
        "الحي",
        "الشارع",
        "ملاحظات",
        "نوع السيارة",
        "اسم المسجّل",
        "موقع الشارع",
    ]
    col_widths = [22, 26, 18, 18, 22, 40, 14, 20, 26]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill
        cell.alignment = ha; cell.border = brd
    ws.row_dimensions[1].height = 30

    valid_rows = []
    for r in rows:
        normalized, ok = normalize_plate_value(full_raw=r.get("full_plate", ""))
        if not ok:
            continue
        rr = dict(r)
        rr["full_plate"] = normalized
        valid_rows.append(rr)

    for i, r in enumerate(valid_rows, 1):
        fill = fe if i % 2 == 0 else fo
        vals = [
            r.get("full_plate", ""),
            r.get("gps", ""),
            r.get("recording_date", ""),
            r.get("district_name", ""),
            r.get("street_name", "غير محدد"),
            r.get("location_details", ""),
            r.get("vehicle_type", "ملاكى"),
            r.get("recorder_name", ""),
            _row_street_location(r),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i + 1, column=col, value=v)
            cell.font = pf if col == 1 else df
            cell.alignment = ca
            cell.border = brd; cell.fill = fill

    for col, w in zip("ABCDEFGHI", col_widths):
        ws.column_dimensions[col].width = w

    content = await workbook_to_bytes_async(wb)
    filename = f"اللوحات_المطابقة_{sheet_name}.xlsx"

    return StreamingResponse(
        io.BytesIO(content),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": _content_disposition(filename, "matched_plates.xlsx")},
    )


def _parse_excel_sync(source: bytes | str) -> tuple[list[dict], int]:
    wb = _open_workbook_readonly(source)
    ws = wb.active
    rows_out: list[dict] = []
    headers: list[str] = []

    for ri, row in enumerate(ws.iter_rows(values_only=True)):
        if ri == 0:
            headers = [str(c).strip() if c else "" for c in row]
            continue
        if all(c is None for c in row):
            continue

        def cell_at(ci: int) -> str:
            if ci < 0 or ci >= len(row) or row[ci] is None:
                return ""
            return str(row[ci]).strip()

        def col(name: str, fallback: int) -> str:
            for i, h in enumerate(headers):
                hs = str(h).strip() if h else ""
                if not hs:
                    continue
                if name == "الشارع":
                    if hs == "موقع الشارع" or hs.startswith("موقع الشارع"):
                        continue
                    if "الشارع" in hs:
                        return cell_at(i)
                    continue
                if name == "الحي":
                    if "الحي" in hs:
                        return cell_at(i)
                    continue
                if name not in hs:
                    continue
                return cell_at(i)
            return cell_at(fallback)

        def col_street_location(fallback: int) -> str:
            for i, h in enumerate(headers):
                hs = str(h).strip() if h else ""
                if hs == "موقع الشارع":
                    return cell_at(i)
            return cell_at(fallback)

        def col_district() -> str:
            for i, h in enumerate(headers):
                hs = str(h).strip() if h else ""
                if "الحي" in hs:
                    return cell_at(i)
            return ""

        def col_by_substrings(substrings: tuple[str, ...], fallback: int) -> str:
            """First header cell containing any substring wins (order matters for legacy vs new names)."""
            for sub in substrings:
                if not sub:
                    continue
                for i, h in enumerate(headers):
                    hs = str(h).strip() if h else ""
                    if not hs:
                        continue
                    if hs == "موقع الشارع" or hs.startswith("موقع الشارع"):
                        continue
                    if sub in hs:
                        return cell_at(i)
            return cell_at(fallback)

        # Fallbacks: new export = plate,gps,date,district,street,... ; legacy (no الحي) matches headers first.
        rows_out.append({
            "full_plate":       col("اللوحة", 0),
            "gps":              col("GPS", 1),
            "recording_date":   col("التسجيل", 2),
            "district_name":    col_district(),
            "street_name":      col("الشارع", 4),
            "location_details": col_by_substrings(
                ("ملاحظات", "تفاصيل الموقع", "الموقع"), 5
            ),
            "vehicle_type":     col_by_substrings(
                ("نوع السيارة", "نوع المركبة", "المركبة"), 6
            ),
            "recorder_name":    col("المسجّل", 7),
            "street_location":  col_street_location(8),
            "notes":            col("ملاحظات", -1),
        })

    return rows_out, len(rows_out)


@router.post("/parse-excel")
async def parse_excel(file: UploadFile = File(...)):
    if not file:
        raise HTTPException(status_code=400, detail="لم يتم رفع ملف")

    tmp_path = None
    try:
        tmp_path = await save_upload_to_temp_with_limit(
            file,
            max_bytes=MAX_EXCEL_BYTES,
            max_mb=30,
            prefix="excel_parse_",
            suffix=".xlsx",
        )
        rows_out, total = await asyncio.to_thread(_parse_excel_sync, tmp_path)

        return JSONResponse({"rows": rows_out, "total": total})

    except Exception:
        # SECURITY FIX: hiding internal exception details from client
        logger.exception("Failed parsing excel upload")
        raise HTTPException(
            status_code=500,
            detail="An internal error occurred. Please try again.",
        )
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass


@router.post("/parse-export-append")
async def parse_export_append(file: UploadFile = File(...)):
    if not file:
        raise HTTPException(status_code=400, detail="لم يتم رفع ملف")
    tmp_path = None
    try:
        tmp_path = await save_upload_to_temp_with_limit(
            file,
            max_bytes=MAX_EXCEL_BYTES,
            max_mb=30,
            prefix="excel_parse_append_",
            suffix=".xlsx",
        )
        rows_out, total = await asyncio.to_thread(_parse_append_excel_sync, tmp_path)
        return JSONResponse({"rows": rows_out, "total": total})
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception:
        logger.exception("Failed parsing append export excel")
        raise HTTPException(
            status_code=500,
            detail="An internal error occurred. Please try again.",
        )
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass


@router.post("/parse-check-session-append")
async def parse_check_session_append(file: UploadFile = File(...)):
    """Append file for check session: 6 columns (matches export-check-session)."""
    if not file:
        raise HTTPException(status_code=400, detail="لم يتم رفع ملف")
    tmp_path = None
    try:
        tmp_path = await save_upload_to_temp_with_limit(
            file,
            max_bytes=MAX_EXCEL_BYTES,
            max_mb=30,
            prefix="excel_parse_check_session_",
            suffix=".xlsx",
        )
        rows_out, total = await asyncio.to_thread(_parse_check_session_append_sync, tmp_path)
        return JSONResponse({"rows": rows_out, "total": total})
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception:
        logger.exception("Failed parsing check-session append excel")
        raise HTTPException(
            status_code=500,
            detail="An internal error occurred. Please try again.",
        )
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass