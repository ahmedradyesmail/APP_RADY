import io
import json
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import APIRouter, Depends, Form, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse
from urllib.parse import quote

from dependencies.auth import get_current_user
from services.plate_utils import (
    normalize_plate,
    auto_detect_plate_col,
    auto_detect_plate_col_from_row3,
)
from services.excel_utils import (
    apply_excel_style,
    find_best_sheet_async,
    load_workbook_from_bytes_async,
    load_workbook_maybe_encrypted_async,
    workbook_to_bytes_async,
)
from services.upload_security import MAX_EXCEL_BYTES, read_upload_with_limit


logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/api",
    tags=["gps"],
    dependencies=[Depends(get_current_user)],
)


def _cell_val(row, ci):
    if ci is None or ci >= len(row):
        return ""
    v = row[ci]
    return str(v).strip() if v is not None else ""


@router.post("/parse-gps-excel")
async def parse_gps_excel(
    file:      UploadFile = File(...),
    label_col: str        = Form(""),
    gps_col:   str        = Form("GPS"),
    label_cols_json: str  = Form(""),
):
    # SECURITY FIX: file size limit to prevent DoS via large uploads
    content = await read_upload_with_limit(file, MAX_EXCEL_BYTES, 30)
    try:
        wb = await load_workbook_from_bytes_async(content)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return JSONResponse(
                {"points": [], "total": 0, "skipped": 0, "headers": [], "label_col_used": ""}
            )

        headers = [str(c).strip() if c is not None else "" for c in rows[0]]

        gps_col = (gps_col or "").strip() or "GPS"
        gps_idx = headers.index(gps_col) if gps_col in headers else None
        if gps_idx is None and "GPS" in headers:
            gps_col = "GPS"
            gps_idx = headers.index("GPS")
        if gps_idx is None and "موقع الشارع" in headers:
            gps_col = "موقع الشارع"
            gps_idx = headers.index("موقع الشارع")
        if gps_idx is None:
            raise HTTPException(
                status_code=400,
                detail="لا يوجد عمود GPS أو موقع الشارع في الملف",
            )

        label_col  = label_col.strip()
        label_idx  = (
            headers.index(label_col)
            if label_col and label_col in headers
            else None
        )

        label_cols: list[str] = []
        if label_cols_json.strip():
            try:
                raw = json.loads(label_cols_json)
                if isinstance(raw, list):
                    label_cols = [str(x).strip() for x in raw if str(x).strip() and str(x).strip() in headers]
            except Exception:
                label_cols = []

        points  = []
        skipped = 0

        for row in rows[1:]:
            if all(c is None for c in row):
                continue
            raw = row[gps_idx] if gps_idx < len(row) else None
            s   = str(raw).strip() if raw is not None else ""
            if not s or "," not in s:
                skipped += 1
                continue
            parts = s.split(",")
            if len(parts) < 2:
                skipped += 1
                continue
            try:
                lat = float(parts[0].strip())
                lng = float(parts[1].strip())
                label_val = (
                    str(row[label_idx]).strip()
                    if label_idx is not None
                    and label_idx < len(row)
                    and row[label_idx] is not None
                    else ""
                )
                fields = {}
                for c in label_cols:
                    ci = headers.index(c) if c in headers else None
                    if ci is not None and ci < len(row) and row[ci] is not None:
                        fields[c] = str(row[ci]).strip()
                    else:
                        fields[c] = ""

                points.append({
                    "lat": lat,
                    "lng": lng,
                    "label": label_val,
                    "fields": fields,
                    "gps_col_used": gps_col,
                })
            except Exception:
                skipped += 1

        return JSONResponse({
            "points":        points,
            "total":         len(points),
            "skipped":       skipped,
            "headers":       [h for h in headers if h],
            "label_col_used": headers[label_idx] if label_idx is not None else "",
            "gps_col_used":  gps_col,
        })

    except HTTPException:
        raise
    except Exception:
        # SECURITY FIX: hiding internal exception details from client
        logger.exception("Failed parsing GPS excel")
        raise HTTPException(
            status_code=500,
            detail="An internal error occurred. Please try again.",
        )


@router.post("/check-gps-data")
async def check_gps_data(
    large_file:  UploadFile = File(...),
    small_file:  UploadFile = File(...),
    password:    str        = Form(""),
    large_col:   str        = Form(""),
    small_col:   str        = Form(""),
    large_sheet: str        = Form(""),
    small_sheet: str        = Form(""),
):
    # SECURITY FIX: file size limit to prevent DoS via large uploads
    lc_bytes = await read_upload_with_limit(large_file, MAX_EXCEL_BYTES, 30)
    # SECURITY FIX: file size limit to prevent DoS via large uploads
    sc_bytes = await read_upload_with_limit(small_file, MAX_EXCEL_BYTES, 30)

    try:
        large_wb = await load_workbook_maybe_encrypted_async(lc_bytes, password.strip())
    except ValueError:
        # SECURITY FIX: hiding internal exception details from client
        logger.exception("Failed to open encrypted large workbook in GPS check")
        raise HTTPException(
            status_code=400,
            detail="An internal error occurred. Please try again.",
        )

    try:
        small_wb = await load_workbook_from_bytes_async(sc_bytes)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"تعذّر فتح الملف الصغير: {e}")

    large_ws = (
        large_wb[large_sheet]
        if large_sheet and large_sheet in large_wb.sheetnames
        else await find_best_sheet_async(large_wb)
    )
    small_ws = (
        small_wb[small_sheet]
        if small_sheet and small_sheet in small_wb.sheetnames
        else await find_best_sheet_async(small_wb)
    )

    ld = list(large_ws.iter_rows(values_only=True))
    if not ld:
        raise HTTPException(status_code=400, detail="الملف الكبير فارغ")

    lh = [str(h).strip() if h is not None else "" for h in ld[0]]
    lc = large_col.strip() or auto_detect_plate_col(lh)

    if not lc or lc not in lh:
        raise HTTPException(
            status_code=422,
            detail=f"لم يُعثر على عمود اللوحة في الملف الكبير. الأعمدة: {lh}",
        )

    gps_col = next((h for h in lh if h.strip() == "GPS"), None)
    if not gps_col:
        raise HTTPException(
            status_code=400, detail="لا يوجد عمود GPS في الملف الكبير"
        )

    date_col  = next((h for h in lh if "تاريخ" in h), None)
    type_col  = next((h for h in lh if "نوع" in h),   None)
    notes_col = next((h for h in lh if "ملاحظات" in h), None)

    lci      = lh.index(lc)
    gps_ci   = lh.index(gps_col)
    date_ci  = lh.index(date_col)  if date_col  else None
    type_ci  = lh.index(type_col)  if type_col  else None
    notes_ci = lh.index(notes_col) if notes_col else None

    lookup = {}
    for row in ld[1:]:
        if all(v is None for v in row):
            continue
        rp   = row[lci] if lci < len(row) else None
        norm = normalize_plate(rp)
        if not norm:
            continue
        lookup.setdefault(norm, []).append({
            "plate":        str(rp or "").strip(),
            "gps":          _cell_val(row, gps_ci),
            "date":         _cell_val(row, date_ci),
            "vehicle_type": _cell_val(row, type_ci),
            "notes":        _cell_val(row, notes_ci),
        })

    sd = list(small_ws.iter_rows(values_only=True))
    if not sd:
        raise HTTPException(status_code=400, detail="الملف الصغير فارغ")

    sh = [str(h).strip() if h is not None else "" for h in sd[0]]
    row3 = sd[2] if len(sd) > 2 else None
    sc = small_col.strip() or auto_detect_plate_col(sh) or auto_detect_plate_col_from_row3(sh, row3)

    if not sc or sc not in sh:
        raise HTTPException(
            status_code=422,
            detail=f"لم يُعثر على عمود اللوحة في الملف الصغير. الأعمدة: {sh}",
        )

    sci     = sh.index(sc)
    matched = []
    seen    = set()

    for row in sd[1:]:
        if all(v is None for v in row):
            continue
        rp   = row[sci] if sci < len(row) else None
        norm = normalize_plate(rp)
        if not norm:
            continue
        if norm in lookup:
            for item in lookup[norm]:
                key = (item["plate"], item["gps"])
                if key not in seen:
                    seen.add(key)
                    matched.append(item)

    if not matched:
        return JSONResponse({"detail": "لا توجد تطابقات بين الملفين", "vehicles": []})

    # Filter rows that have valid GPS coordinates
    valid = []
    for item in matched:
        gps = item.get("gps", "")
        if not gps or gps in ("None", "") or "," not in gps:
            continue
        parts = gps.split(",")
        try:
            float(parts[0].strip())
            float(parts[1].strip())
            valid.append(item)
        except Exception:
            continue

    return JSONResponse({
        "vehicles": valid,
        "total":    len(valid),
        "skipped":  len(matched) - len(valid),
    })


@router.post("/parse-ref-plates")
async def parse_ref_plates(
    file: UploadFile = File(...),
    col:  str        = Form(""),
):
    # SECURITY FIX: file size limit to prevent DoS via large uploads
    content = await read_upload_with_limit(file, MAX_EXCEL_BYTES, 30)
    try:
        wb = await load_workbook_from_bytes_async(content)
        ws = await find_best_sheet_async(wb)
        rows    = list(ws.iter_rows(values_only=True))
        if not rows:
            return JSONResponse({"plates": [], "total": 0, "col_used": ""})

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        col     = col.strip()
        col_idx = None

        if col and col in headers:
            col_idx = headers.index(col)
        if col_idx is None:
            detected = auto_detect_plate_col(headers)
            if detected and detected in headers:
                col_idx = headers.index(detected)
        if col_idx is None:
            return JSONResponse(
                status_code=422,
                content={
                    "detail":  f"لم يُعثر على عمود اللوحة. الأعمدة: {headers}",
                    "headers": headers,
                    "code":    "COL_NOT_FOUND",
                },
            )

        plates = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None and str(val).strip():
                plates.append(str(val).strip())

        return JSONResponse({
            "plates":   plates,
            "total":    len(plates),
            "col_used": headers[col_idx],
        })

    except Exception:
        # SECURITY FIX: hiding internal exception details from client
        logger.exception("Failed parsing reference plates")
        raise HTTPException(
            status_code=500,
            detail="An internal error occurred. Please try again.",
        )


@router.post("/check-ref-plate")
async def check_ref_plate(
    file: UploadFile = File(...),
    plate: str = Form(""),
    col: str = Form(""),
):
    raw_plate = (plate or "").strip()
    norm_target = normalize_plate(raw_plate)
    if not norm_target:
        raise HTTPException(status_code=400, detail="أدخل رقم لوحة صحيح")

    # SECURITY FIX: file size limit to prevent DoS via large uploads
    content = await read_upload_with_limit(file, MAX_EXCEL_BYTES, 30)
    try:
        wb = await load_workbook_from_bytes_async(content)
        ws = await find_best_sheet_async(wb)
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return JSONResponse({
                "exists": False,
                "matched_plate": "",
                "total_scanned": 0,
                "col_used": "",
            })

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        selected_col = (col or "").strip()
        col_idx = None

        if selected_col and selected_col in headers:
            col_idx = headers.index(selected_col)
        if col_idx is None:
            detected = auto_detect_plate_col(headers)
            if detected and detected in headers:
                col_idx = headers.index(detected)
        if col_idx is None:
            return JSONResponse(
                status_code=422,
                content={
                    "detail": f"لم يُعثر على عمود اللوحة. الأعمدة: {headers}",
                    "headers": headers,
                    "code": "COL_NOT_FOUND",
                },
            )

        total_scanned = 0
        matched_plate = ""
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            val = row[col_idx] if col_idx < len(row) else None
            if val is None:
                continue
            raw_val = str(val).strip()
            if not raw_val:
                continue
            total_scanned += 1
            if normalize_plate(raw_val) == norm_target:
                matched_plate = raw_val
                break

        return JSONResponse({
            "exists": bool(matched_plate),
            "matched_plate": matched_plate,
            "total_scanned": total_scanned,
            "col_used": headers[col_idx],
        })
    except HTTPException:
        raise
    except Exception:
        # SECURITY FIX: hiding internal exception details from client
        logger.exception("Failed checking single reference plate")
        raise HTTPException(
            status_code=500,
            detail="An internal error occurred. Please try again.",
        )


@router.post("/export-gps-excel")
async def export_gps_excel(
    results_json: str = Form("[]"),
    failed_json:  str = Form("[]"),
    my_lat:       str = Form(""),
    my_lon:       str = Form(""),
):
    try:
        results = json.loads(results_json)
        failed  = json.loads(failed_json)
    except Exception:
        raise HTTPException(status_code=400, detail="تنسيق JSON خاطئ")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "أقرب المركبات"
    ws.sheet_view.rightToLeft = True

    hf     = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    hfill  = PatternFill("solid", start_color="0D6B5E")
    ha     = Alignment(horizontal="center", vertical="center")
    ca     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    df     = Font(name="Arial", size=11)
    lf_map = Font(name="Arial", size=11, color="0563C1", underline="single")
    thin   = Side(style="thin", color="BFBFBF")
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    fe     = PatternFill("solid", start_color="E0F2F1")
    fo     = PatternFill("solid", start_color="FFFFFF")
    rank1f = PatternFill("solid", start_color="C8E6C9")

    headers    = ["#", "رقم اللوحة", "الخريطة", "النوع", "ملاحظات",
                  "المسافة (km)", "الوقت (دقيقة)", "تاريخ التسجيل"]
    col_widths = [5, 22, 28, 14, 22, 14, 14, 20]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill
        cell.alignment = ha; cell.border = brd
    ws.row_dimensions[1].height = 30

    for i, r in enumerate(results, 1):
        fill     = rank1f if r.get("rank") == 1 else (fe if i % 2 == 0 else fo)
        gps      = r.get("gps", "")
        maps_url = ""

        if gps and "," in gps and my_lat and my_lon:
            parts = gps.split(",")
            try:
                vlat     = float(parts[0].strip())
                vlng     = float(parts[1].strip())
                maps_url = (
                    f"https://www.google.com/maps/dir/{my_lat},{my_lon}/{vlat},{vlng}"
                )
            except Exception:
                pass

        row_vals = [
            r.get("rank", i),
            r.get("plate", ""),
            maps_url or "",
            r.get("vehicle_type", ""),
            r.get("notes", ""),
            r.get("distance_km", ""),
            r.get("duration_min", ""),
            r.get("date", ""),
        ]

        for col, v in enumerate(row_vals, 1):
            cell = ws.cell(row=i + 1, column=col, value=v)
            cell.border = brd; cell.fill = fill
            if col == 3 and maps_url:
                cell.value     = "📍 فتح الخريطة"
                cell.hyperlink = maps_url
                cell.font      = lf_map
                cell.alignment = ca
            else:
                cell.font      = df
                cell.alignment = ca

    for idx, w in enumerate(col_widths):
        ws.column_dimensions[chr(65 + idx)].width = w

    if failed:
        ws_f = wb.create_sheet("فشلت")
        apply_excel_style(
            ws_f,
            ["رقم اللوحة", "GPS", "السبب"],
            [
                {
                    "رقم اللوحة": f.get("plate", ""),
                    "GPS":         f.get("gps",   ""),
                    "السبب":       f.get("reason", ""),
                }
                for f in failed
            ],
        )

    ws_s = wb.create_sheet("ملخص")
    apply_excel_style(ws_s, ["البند", "القيمة"], [
        {"البند": "موقع المستخدم",          "القيمة": f"{my_lat}, {my_lon}"},
        {"البند": "إجمالي مركبات مطابَقة",  "القيمة": len(results)},
        {"البند": "مركبات فشلت",             "القيمة": len(failed)},
        {"البند": "أقرب مركبة",              "القيمة": results[0]["plate"]        if results else "—"},
        {"البند": "أقل مسافة (km)",          "القيمة": results[0]["distance_km"]  if results else "—"},
        {"البند": "أقل وقت (دقيقة)",         "القيمة": results[0]["duration_min"] if results else "—"},
    ])

    content = await workbook_to_bytes_async(wb)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"أقرب_المركبات_{ts}.xlsx"
    encoded_filename = quote(filename, safe="")

    return StreamingResponse(
        io.BytesIO(content),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                f'attachment; filename="nearest_vehicles_{ts}.xlsx"; '
                f"filename*=UTF-8''{encoded_filename}"
            )
        },
    )