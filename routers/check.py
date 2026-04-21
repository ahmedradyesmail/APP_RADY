import asyncio
import base64
import io
import json
import logging
import os
import tempfile
import time
import re
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, JSONResponse
import openpyxl
from sqlalchemy.orm import Session

from db import get_db
from dependencies.auth import get_current_user
from models import UserGroup
from models.user import User
from config import settings
from services.check_match import run_check_plates_sync
from services.check_postgres import (
    CHECK_PG_MAX_LARGE_BYTES,
    CHECK_PG_MAX_ROWS_PER_USER,
    collect_gps_vehicles_stored_sync,
    delete_import_sync,
    get_stored_large_meta_for_check_sync,
    import_large_workbook_sync,
    list_imports_sync,
    run_check_plates_postgres_sync,
)
from services.check_temp_storage import (
    CHECK_TEMP_MAX_LARGE_BYTES,
    CHECK_TEMP_TTL_MINUTES,
    delete_temp_session_sync,
    ping_temp_session_sync,
    purge_expired_temp_sessions_sync,
    query_temp_plates_sync,
    start_temp_session_sync,
    upload_large_temp_plates_sync,
)
from services.check_queue import (
    CheckQueueFullError,
    enqueue_check_job,
    queue_depth,
)
from services.excel_utils import (
    find_best_sheet_async,
    load_workbook_from_bytes_async,
    load_workbook_maybe_encrypted_async,
)
from services.job_store import (
    TTL_PROCESSING_SEC,
    TTL_TERMINAL_SEC,
    job_get,
    job_save,
    new_job_id,
    schedule_job_cleanup,
)
from services.plate_utils import auto_detect_plate_col, auto_detect_plate_col_from_row3
from services.upload_security import MAX_EXCEL_BYTES, read_upload_with_limit


logger = logging.getLogger(__name__)
_JOB_ID_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$",
    re.IGNORECASE,
)

_RATE_LIMIT_LOCK = asyncio.Lock()
_USER_CHECK_TS: dict[int, list[float]] = {}
_RATE_WINDOW_SEC = 60.0

router = APIRouter(
    prefix="/api",
    tags=["check"],
    dependencies=[Depends(get_current_user)],
)


def _check_pg_dsn() -> str | None:
    u = (settings.check_postgres_dsn or "").strip()
    return u or None


def _job_id_valid(job_id: str) -> bool:
    return bool(_JOB_ID_RE.match((job_id or "").strip()))


def _group_context(db: Session, user: User) -> tuple[str | None, bool]:
    u = db.get(User, user.id)
    if not u or not getattr(u, "group_id", None):
        return None, False
    g = db.get(UserGroup, u.group_id)
    return (g.name if g else None), True


def _group_rows_limit_for_user(db: Session, user: User) -> int | None:
    u = db.get(User, user.id)
    if not u or not getattr(u, "group_id", None):
        return None
    g = db.get(UserGroup, u.group_id)
    if not g:
        return None
    try:
        v = int(getattr(g, "max_stored_large_rows", 0) or 0)
    except Exception:
        return None
    return v if v > 0 else None


def _user_rows_limit_for_user(db: Session, user: User) -> int | None:
    u = db.get(User, user.id)
    if not u:
        return None
    try:
        v = int(getattr(u, "max_stored_large_rows", 0) or 0)
    except Exception:
        return None
    return v if v > 0 else None


def _form_bool(raw: str) -> bool:
    return str(raw or "").strip().lower() in ("1", "true", "yes", "on")


def _get_headers(ws) -> list[str]:
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        return [str(h).strip() if h is not None else "" for h in row]
    return []


def _get_row_values(ws, row_no: int) -> tuple | None:
    for row in ws.iter_rows(min_row=row_no, max_row=row_no, values_only=True):
        return row
    return None


def _collect_column_samples(
    ws,
    headers: list[str],
    *,
    min_row: int = 2,
    max_rows: int = 25,
) -> dict[str, list[str]]:
    out: dict[str, list[str]] = {h: [] for h in headers if h}
    if not out:
        return out
    taken = 0
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        if taken >= max_rows:
            break
        taken += 1
        for i, h in enumerate(headers):
            if not h:
                continue
            v = row[i] if i < len(row) else None
            if v is None:
                continue
            sv = str(v).strip()
            if sv:
                out[h].append(sv)
    return out


@router.post("/check-headers")
async def check_headers(
    large_file: UploadFile | None = File(None),
    small_file: UploadFile | None = File(None),
    password: str = Form(""),
):
    result = {}

    if large_file:
        try:
            # SECURITY FIX: file size limit to prevent DoS via large uploads
            content = await read_upload_with_limit(large_file, MAX_EXCEL_BYTES, 30)
            wb = await load_workbook_maybe_encrypted_async(content, password.strip())
            ws = await find_best_sheet_async(wb)
            headers = _get_headers(ws)
            result["large"] = {
                "headers": headers,
                "detected": auto_detect_plate_col(headers),
                "sheet_name": ws.title,
                "all_sheets": wb.sheetnames,
            }
        except Exception as e:
            result["large"] = {"error": str(e)}

    if small_file:
        try:
            # SECURITY FIX: file size limit to prevent DoS via large uploads
            content = await read_upload_with_limit(small_file, MAX_EXCEL_BYTES, 30)
            wb = await load_workbook_from_bytes_async(content)
            ws = await find_best_sheet_async(wb)
            headers = _get_headers(ws)
            row3 = _get_row_values(ws, 3)
            col_samples = _collect_column_samples(ws, headers)
            result["small"] = {
                "headers": headers,
                "detected": auto_detect_plate_col(headers)
                or auto_detect_plate_col_from_row3(headers, row3),
                "sheet_name": ws.title,
                "all_sheets": wb.sheetnames,
                "column_samples": col_samples,
            }
        except Exception as e:
            result["small"] = {"error": str(e)}

    return JSONResponse(result)


@router.get("/check/capabilities")
async def check_capabilities():
    return JSONResponse(
        {
            "postgres_large_enabled": _check_pg_dsn() is not None,
            "max_large_import_mb": 15,
            "max_rows_per_user_stored": CHECK_PG_MAX_ROWS_PER_USER,
            "check_temp_ttl_minutes": CHECK_TEMP_TTL_MINUTES,
        }
    )


@router.post("/check/temp/session/start")
async def check_temp_session_start(current_user: User = Depends(get_current_user)):
    dsn = _check_pg_dsn()
    if not dsn:
        raise HTTPException(status_code=503, detail="CHECK_POSTGRES_URL is not configured.")
    try:
        await asyncio.to_thread(purge_expired_temp_sessions_sync, dsn)
        token = await asyncio.to_thread(
            start_temp_session_sync,
            dsn,
            current_user.id,
            current_user.is_admin,
        )
    except Exception:
        logger.exception("start_temp_session_sync failed")
        raise HTTPException(status_code=500, detail="Failed to start temp check session.")
    return JSONResponse({"session_token": token, "ttl_minutes": CHECK_TEMP_TTL_MINUTES})


@router.post("/check/temp/session/ping")
async def check_temp_session_ping(
    session_token: str = Form(""),
    current_user: User = Depends(get_current_user),
):
    dsn = _check_pg_dsn()
    if not dsn:
        raise HTTPException(status_code=503, detail="CHECK_POSTGRES_URL is not configured.")
    token = (session_token or "").strip()
    if not token:
        raise HTTPException(status_code=400, detail="session_token is required.")
    try:
        await asyncio.to_thread(purge_expired_temp_sessions_sync, dsn)
        ok = await asyncio.to_thread(
            ping_temp_session_sync,
            dsn,
            current_user.id,
            current_user.is_admin,
            token,
        )
    except Exception:
        logger.exception("ping_temp_session_sync failed")
        raise HTTPException(status_code=500, detail="Failed to ping temp session.")
    if not ok:
        raise HTTPException(status_code=404, detail="Temp session not found or expired.")
    return JSONResponse({"ok": True})


@router.post("/check/temp/session/close")
async def check_temp_session_close(
    session_token: str = Form(""),
    current_user: User = Depends(get_current_user),
):
    dsn = _check_pg_dsn()
    if not dsn:
        raise HTTPException(status_code=503, detail="CHECK_POSTGRES_URL is not configured.")
    token = (session_token or "").strip()
    if not token:
        raise HTTPException(status_code=400, detail="session_token is required.")
    try:
        closed = await asyncio.to_thread(
            delete_temp_session_sync,
            dsn,
            current_user.id,
            current_user.is_admin,
            token,
        )
    except Exception:
        logger.exception("delete_temp_session_sync failed")
        raise HTTPException(status_code=500, detail="Failed to close temp session.")
    return JSONResponse({"ok": bool(closed)})


@router.post("/check/temp/upload-large")
async def check_temp_upload_large(
    large_file: UploadFile = File(...),
    session_token: str = Form(""),
    password: str = Form(""),
    large_col: str = Form(""),
    large_sheet: str = Form(""),
    current_user: User = Depends(get_current_user),
):
    dsn = _check_pg_dsn()
    if not dsn:
        raise HTTPException(status_code=503, detail="CHECK_POSTGRES_URL is not configured.")
    token = (session_token or "").strip()
    if not token:
        raise HTTPException(status_code=400, detail="session_token is required.")
    try:
        await asyncio.to_thread(purge_expired_temp_sessions_sync, dsn)
        content = await read_upload_with_limit(large_file, CHECK_TEMP_MAX_LARGE_BYTES, 15)
        out = await asyncio.to_thread(
            upload_large_temp_plates_sync,
            dsn,
            current_user.id,
            current_user.is_admin,
            session_token=token,
            large_bytes=content,
            password=password.strip(),
            large_col=large_col.strip(),
            large_sheet=large_sheet.strip(),
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    except HTTPException:
        raise
    except Exception:
        logger.exception("upload_large_temp_plates_sync failed")
        raise HTTPException(status_code=500, detail="Failed to upload temp large file.")
    return JSONResponse({"ok": True, **out})


@router.post("/check/temp/query")
async def check_temp_query(
    session_token: str = Form(""),
    plates_text: str = Form(""),
    current_user: User = Depends(get_current_user),
):
    dsn = _check_pg_dsn()
    if not dsn:
        raise HTTPException(status_code=503, detail="CHECK_POSTGRES_URL is not configured.")
    token = (session_token or "").strip()
    if not token:
        raise HTTPException(status_code=400, detail="session_token is required.")
    try:
        await asyncio.to_thread(purge_expired_temp_sessions_sync, dsn)
        out = await asyncio.to_thread(
            query_temp_plates_sync,
            dsn,
            current_user.id,
            current_user.is_admin,
            session_token=token,
            plates_text=plates_text,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    except Exception:
        logger.exception("query_temp_plates_sync failed")
        raise HTTPException(status_code=500, detail="Failed to check plates.")
    return JSONResponse(out)


@router.get("/check/stored-large-meta")
async def check_stored_large_meta(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    dsn = _check_pg_dsn()
    if not dsn:
        raise HTTPException(
            status_code=503,
            detail="CHECK_POSTGRES_URL is not configured.",
        )
    gn, ig = _group_context(db, current_user)
    try:
        meta = await asyncio.to_thread(
            get_stored_large_meta_for_check_sync,
            dsn,
            current_user.id,
            current_user.is_admin,
        )
    except Exception:
        logger.exception("get_stored_large_meta_for_check_sync failed")
        raise HTTPException(
            status_code=500,
            detail="Failed to read stored large file metadata.",
        )
    if not meta:
        return JSONResponse(
            {
                "has_data": False,
                "headers": [],
                "sheet_name": "",
                "row_count": 0,
                "updated_at": None,
                "imports": [],
                "group_name": gn,
                "in_group": ig,
            }
        )
    return JSONResponse(
        {"has_data": True, **meta, "group_name": gn, "in_group": ig}
    )


@router.get("/check/stored-imports")
async def check_stored_imports_list(
    current_user: User = Depends(get_current_user),
):
    dsn = _check_pg_dsn()
    if not dsn:
        raise HTTPException(
            status_code=503,
            detail="CHECK_POSTGRES_URL is not configured.",
        )
    try:
        imports = await asyncio.to_thread(
            list_imports_sync, dsn, current_user.id, current_user.is_admin
        )
    except Exception:
        logger.exception("list_imports_sync failed")
        raise HTTPException(
            status_code=500,
            detail="Failed to list stored imports.",
        )
    return JSONResponse({"imports": imports})


@router.delete("/check/stored-imports/{import_id}")
async def check_stored_import_delete(
    import_id: int,
    current_user: User = Depends(get_current_user),
):
    dsn = _check_pg_dsn()
    if not dsn:
        raise HTTPException(
            status_code=503,
            detail="CHECK_POSTGRES_URL is not configured.",
        )
    try:
        ok = await asyncio.to_thread(
            delete_import_sync,
            dsn,
            current_user.id,
            current_user.is_admin,
            import_id,
        )
    except Exception:
        logger.exception("delete_import_sync failed")
        raise HTTPException(
            status_code=500,
            detail="Failed to delete import.",
        )
    if not ok:
        raise HTTPException(status_code=404, detail="Import not found.")
    return JSONResponse({"ok": True})


@router.post("/check/import-large")
async def check_import_large(
    large_file: UploadFile = File(...),
    password: str = Form(""),
    large_col: str = Form(""),
    large_sheet: str = Form(""),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    dsn = _check_pg_dsn()
    if not dsn:
        raise HTTPException(
            status_code=503,
            detail="CHECK_POSTGRES_URL is not configured.",
        )
    await _rate_limit_user_check(current_user.id)
    try:
        content = await read_upload_with_limit(large_file, CHECK_PG_MAX_LARGE_BYTES, 15)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    try:
        grp_limit = _group_rows_limit_for_user(db, current_user)
        usr_limit = _user_rows_limit_for_user(db, current_user)
        summary = await asyncio.to_thread(
            import_large_workbook_sync,
            dsn,
            current_user.id,
            current_user.is_admin,
            content,
            password.strip(),
            large_col.strip(),
            large_sheet.strip(),
            large_file.filename or "upload.xlsx",
            grp_limit,
            usr_limit,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    except Exception:
        logger.exception("import_large_workbook_sync failed")
        raise HTTPException(
            status_code=500,
            detail="An internal error occurred while importing the large file.",
        )
    return JSONResponse({"ok": True, **summary})


@router.post("/check/gps-stored")
async def check_gps_stored(
    small_file: UploadFile = File(...),
    small_col: str = Form(""),
    small_sheet: str = Form(""),
    current_user: User = Depends(get_current_user),
):
    """GPS nearest flow using large rows in Postgres (same JSON shape as /api/check-gps-data)."""
    dsn = _check_pg_dsn()
    if not dsn:
        raise HTTPException(
            status_code=503,
            detail="CHECK_POSTGRES_URL is not configured.",
        )
    sc_bytes = await read_upload_with_limit(small_file, MAX_EXCEL_BYTES, 30)
    try:
        data = await asyncio.to_thread(
            collect_gps_vehicles_stored_sync,
            dsn,
            current_user.id,
            current_user.is_admin,
            sc_bytes,
            small_col.strip(),
            small_sheet.strip(),
        )
    except Exception:
        logger.exception("collect_gps_vehicles_stored_sync failed")
        raise HTTPException(
            status_code=500,
            detail="An internal error occurred.",
        )
    return JSONResponse(data)


def _parse_col_list(raw: str) -> list[str]:
    s = (raw or "").strip()
    if not s:
        return []
    try:
        data = json.loads(s)
        if isinstance(data, list):
            return [str(x).strip() for x in data if str(x).strip()]
    except Exception:
        pass
    return []


def _normalize_small_plates_text(raw: str) -> list[str]:
    lines = []
    for line in (raw or "").splitlines():
        v = " ".join(str(line).strip().split())
        if v:
            lines.append(v)
    return lines


def _plates_text_to_small_xlsx_bytes(raw: str) -> tuple[bytes, str, int]:
    plates = _normalize_small_plates_text(raw)
    if not plates:
        raise HTTPException(status_code=400, detail="يرجى إدخال لوحة واحدة على الأقل في خانة اللوحات النصية.")
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    header = "رقم اللوحة"
    ws.title = "لوحات"
    ws.append([header])
    for plate in plates:
        ws.append([plate])
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.read(), header, len(plates)


async def _rate_limit_user_check(user_id: int) -> None:
    limit_per_min = max(1, int(settings.check_user_rate_limit_per_minute))
    now = time.time()
    async with _RATE_LIMIT_LOCK:
        ts = _USER_CHECK_TS.setdefault(user_id, [])
        cutoff = now - _RATE_WINDOW_SEC
        while ts and ts[0] < cutoff:
            ts.pop(0)
        if len(ts) >= limit_per_min:
            raise HTTPException(
                status_code=429,
                detail=f"Too many check jobs for this user. Max {limit_per_min}/minute.",
            )
        ts.append(now)


async def _save_upload_to_temp_with_limit(
    upload: UploadFile,
    *,
    max_bytes: int,
    max_mb: int,
    prefix: str,
) -> str:
    fd, tmp_path = tempfile.mkstemp(prefix=prefix, suffix=".xlsx")
    os.close(fd)
    written = 0
    try:
        with open(tmp_path, "wb") as out:
            while True:
                chunk = await upload.read(1024 * 1024)
                if not chunk:
                    break
                written += len(chunk)
                if written > max_bytes:
                    raise HTTPException(
                        status_code=413,
                        detail=f"File too large. Maximum allowed size is {max_mb} MB.",
                    )
                out.write(chunk)
        return tmp_path
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise
    finally:
        await upload.close()


async def process_check_queue_item(item: dict) -> None:
    job_id = item["job_id"]
    mode = item.get("mode") or "files"
    large_path = item.get("large_path")
    small_path = item["small_path"]
    try:
        sc_bytes = await asyncio.to_thread(Path(small_path).read_bytes)
        if mode == "postgres":
            dsn = _check_pg_dsn()
            if not dsn:
                await job_save(
                    job_id,
                    {
                        "status": "error",
                        "data": None,
                        "detail": "Postgres check storage is not configured.",
                    },
                    ttl_seconds=TTL_TERMINAL_SEC,
                )
                return
            raw = await asyncio.to_thread(
                run_check_plates_postgres_sync,
                dsn,
                int(item["user_id"]),
                bool(item.get("is_admin")),
                sc_bytes,
                item.get("password") or "",
                item["small_col"],
                item["small_sheet"],
                item["large_export_cols"],
                item["small_export_cols"],
            )
        else:
            if not large_path:
                await job_save(
                    job_id,
                    {
                        "status": "error",
                        "data": None,
                        "detail": "Missing large file path.",
                    },
                    ttl_seconds=TTL_TERMINAL_SEC,
                )
                return
            lc_bytes = await asyncio.to_thread(Path(large_path).read_bytes)
            raw = await asyncio.to_thread(
                run_check_plates_sync,
                lc_bytes,
                sc_bytes,
                item["password"],
                item["large_col"],
                item["small_col"],
                item["large_sheet"],
                item["small_sheet"],
                item["large_export_cols"],
                item["small_export_cols"],
            )
        if raw["kind"] == "xlsx":
            data_out: dict = {
                "kind": "xlsx",
                "filename": raw["filename"],
                "storage": "inline",
                "content_b64": base64.b64encode(raw["content"]).decode("ascii"),
            }
            if raw.get("preview"):
                data_out["preview"] = raw["preview"]
            await job_save(
                job_id,
                {
                    "status": "done",
                    "data": data_out,
                },
                ttl_seconds=TTL_TERMINAL_SEC,
            )
        else:
            await job_save(
                job_id,
                {
                    "status": "done",
                    "data": {
                        "kind": "json",
                        "status_code": raw["status_code"],
                        "body": raw["body"],
                    },
                },
                ttl_seconds=TTL_TERMINAL_SEC,
            )
    except ValueError as e:
        msg = str(e)
        if "فشل فك تشفير" in msg or "فشل فك" in msg:
            logger.exception("Failed to open encrypted large workbook")
            await job_save(
                job_id,
                {
                    "status": "error",
                    "data": None,
                    "detail": "An internal error occurred. Please try again.",
                },
                ttl_seconds=TTL_TERMINAL_SEC,
            )
        else:
            await job_save(
                job_id,
                {
                    "status": "error",
                    "data": None,
                    "detail": msg,
                },
                ttl_seconds=TTL_TERMINAL_SEC,
            )
    except Exception:
        logger.exception("Check plates job failed")
        await job_save(
            job_id,
            {
                "status": "error",
                "data": None,
                "detail": "An internal error occurred. Please try again.",
            },
            ttl_seconds=TTL_TERMINAL_SEC,
        )
    finally:
        for p in (large_path, small_path):
            if not p:
                continue
            try:
                os.unlink(p)
            except OSError:
                pass
        schedule_job_cleanup(job_id)


@router.post("/check")
async def check_plates(
    large_file: UploadFile | None = File(None),
    small_file: UploadFile | None = File(None),
    small_plates_text: str = Form(""),
    password: str = Form(""),
    large_col: str = Form(""),
    small_col: str = Form(""),
    large_sheet: str = Form(""),
    small_sheet: str = Form(""),
    large_export_cols_json: str = Form(""),
    small_export_cols_json: str = Form(""),
    use_stored_large: str = Form("false"),
    current_user: User = Depends(get_current_user),
):
    await _rate_limit_user_check(current_user.id)
    if await queue_depth() >= max(1, int(settings.check_queue_max_depth)):
        raise HTTPException(
            status_code=429,
            detail="Check queue is busy. Please try again shortly.",
        )

    dsn = _check_pg_dsn()
    use_pg = _form_bool(use_stored_large) and dsn is not None
    if use_pg:
        if large_file is not None and getattr(large_file, "filename", None):
            fn = str(large_file.filename or "").strip()
            if fn:
                raise HTTPException(
                    status_code=400,
                    detail="When use_stored_large is true, do not upload a large file.",
                )
        large_path = None
    else:
        if large_file is None or not str(getattr(large_file, "filename", "") or "").strip():
            raise HTTPException(
                status_code=400,
                detail="large_file is required unless use_stored_large is enabled and configured.",
            )
        large_path = await _save_upload_to_temp_with_limit(
            large_file, max_bytes=MAX_EXCEL_BYTES, max_mb=30, prefix="check_large_"
        )

    use_small_text = bool(_normalize_small_plates_text(small_plates_text))
    if use_small_text:
        try:
            small_bytes, default_small_col, _small_count = _plates_text_to_small_xlsx_bytes(
                small_plates_text
            )
            fd, small_path = tempfile.mkstemp(prefix="check_small_text_", suffix=".xlsx")
            os.close(fd)
            await asyncio.to_thread(Path(small_path).write_bytes, small_bytes)
            if not small_col.strip():
                small_col = default_small_col
        except Exception:
            if large_path:
                try:
                    os.unlink(large_path)
                except OSError:
                    pass
            raise
    else:
        if small_file is None or not str(getattr(small_file, "filename", "") or "").strip():
            if large_path:
                try:
                    os.unlink(large_path)
                except OSError:
                    pass
            raise HTTPException(
                status_code=400,
                detail="small_file مطلوب أو أدخل اللوحات النصية.",
            )
        try:
            small_path = await _save_upload_to_temp_with_limit(
                small_file, max_bytes=MAX_EXCEL_BYTES, max_mb=30, prefix="check_small_"
            )
        except Exception:
            if large_path:
                try:
                    os.unlink(large_path)
                except OSError:
                    pass
            raise

    job_id = new_job_id()
    await job_save(
        job_id,
        {"status": "processing", "data": None},
        ttl_seconds=TTL_PROCESSING_SEC,
    )
    item = {
        "job_id": job_id,
        "mode": "postgres" if use_pg else "files",
        "large_path": large_path,
        "small_path": small_path,
        "password": password.strip(),
        "large_col": large_col.strip(),
        "small_col": small_col.strip(),
        "large_sheet": large_sheet.strip(),
        "small_sheet": small_sheet.strip(),
        "large_export_cols": _parse_col_list(large_export_cols_json),
        "small_export_cols": _parse_col_list(small_export_cols_json),
        "user_id": current_user.id,
        "is_admin": bool(current_user.is_admin),
    }
    try:
        await enqueue_check_job(item)
    except CheckQueueFullError:
        await job_save(
            job_id,
            {"status": "error", "data": None, "detail": "Check queue is busy."},
            ttl_seconds=TTL_TERMINAL_SEC,
        )
        for p in (large_path, small_path):
            if not p:
                continue
            try:
                os.unlink(p)
            except OSError:
                pass
        raise HTTPException(status_code=429, detail="Check queue is busy. Please try again shortly.")
    return JSONResponse({"job_id": job_id, "status": "processing"})


@router.get("/check/status/{job_id}")
async def check_status(job_id: str):
    row = await job_get(job_id)
    if not row:
        raise HTTPException(status_code=404, detail="Job not found")
    return JSONResponse(row)


@router.get("/check/result/{job_id}")
async def check_download_result(job_id: str):
    """Download فرز Excel from inline job payload (no server-side result files)."""
    if not _job_id_valid(job_id):
        raise HTTPException(status_code=400, detail="Invalid job id")
    row = await job_get(job_id)
    if not row:
        raise HTTPException(status_code=404, detail="Job not found")
    if row.get("status") != "done":
        raise HTTPException(status_code=409, detail="Result not ready")
    data = row.get("data") or {}
    if data.get("kind") != "xlsx":
        raise HTTPException(status_code=404, detail="No downloadable file for this job")
    b64 = str(data.get("content_b64") or "")
    if not b64:
        raise HTTPException(status_code=404, detail="Result content missing")
    try:
        content = base64.b64decode(b64, validate=True)
    except Exception:
        raise HTTPException(status_code=500, detail="Result content corrupted")
    fname = data.get("filename") or "التطابقات.xlsx"
    return JSONResponse(
        {
            "filename": fname,
            "content_b64": base64.b64encode(content).decode("ascii"),
        }
    )
